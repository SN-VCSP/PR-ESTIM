
from math import hypot
from typing import List, Tuple, Dict, Optional, Any
import numpy as np
import pandas as pd
import streamlit as st
import folium
from folium.plugins import Draw
from folium.plugins import MarkerCluster  # ⚡ FLUIDITÉ: clustering optionnel des points PR
from folium.features import DivIcon
from streamlit_folium import st_folium
from pyproj import Transformer
from PIL import Image
import sys, os
import inspect
from folium import IFrame

def _round_df0(df: pd.DataFrame, exclude: list[str] | None = None) -> pd.DataFrame:
    """Arrondit à 0 décimal toutes les colonnes numériques, sauf celles de exclude."""
    df2 = df.copy()
    num_cols = df2.select_dtypes(include=[np.number]).columns.tolist()
    if exclude:
        num_cols = [c for c in num_cols if c not in set(exclude)]
    if num_cols:
        df2[num_cols] = df2[num_cols].round(0)
    return df2

# Colonnes à préserver avec leurs décimales dans les tableaux (densité, prix, épaisseur, largeur…)
_DEC_EXCLUDE = [
    "densité_t_m3", "prix_eur_t", "épaisseur_cm", "largeur_m",
    "largeur_équivalente_m", "comptage_équivalent", "hauteur_cm", "rabot_h_cm",
]

    
# =========================
# Configuration et constantes
# =========================
st.set_page_config(page_title="Estimation de surfaces -Sn", layout="wide")


def resource_path(relative_path):
    """Retourne le chemin absolu, compatible PyInstaller"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

logo_path = resource_path("mon_logo1.png")
# 🧭 UX: logo optionnel — ne JAMAIS bloquer le démarrage si 'mon_logo1.png' est absent.
if os.path.exists(logo_path):
    try:
        # Appelle st.logo de manière compatible avec toutes versions
        if "size" in inspect.signature(st.logo).parameters:
            st.logo(logo_path, size="large")
        else:
            st.logo(logo_path)
    except Exception:
        pass  # un logo illisible ne doit pas empêcher l'application de tourner


DEFAULT_WIDTHS: Dict[str, float] = {
    "BAU": 2.5,
    "BDG": 1.0,
    "VL": 3.5,
    "VR": 3.5,
    "VM": 3.5,
    "VS": 3.5,
    "BRET": 3.5,
    "Accès": 5.0,           # Corps d’accès
    "AccoD": 0.30,    # Accotement droit
    "AccoG": 0.30,    # Accotement gauche
}
ALL_ELEMENTS = ["BAU", "BDG", "VL", "VR", "VM", "VS", "BRET", "Accès", "AccoD", "AccoG"]

# Profils (comptages "équivalents" par élément)
PROFILES: Dict[str, Dict[str, float]] = {
    "2_voies": {"BDG": 1, "VR": 1, "VL": 1, "BAU": 1},
    "2_voies_bretelle": {"BDG": 1, "VR": 1, "VL": 1, "BRET": 1, "BAU": 1},
    "3_voies": {"BDG": 1, "VR": 1, "VM": 1, "VL": 1, "BAU": 1},
    "3_voies_bretelle": {"BDG": 1, "VR": 1, "VM": 1, "VL": 1, "BRET": 1, "BAU": 1},
    "4_voies": {"BDG": 1, "VR": 1, "VM": 1, "VL": 1, "VS": 1, "BAU": 1},
    "4_voies_bretelle": {"BDG": 1, "VR": 1, "VM": 1, "VL": 1, "VS": 1, "BRET": 1, "BAU": 1},
    "Accès": {"Accès": 1, "AccoD": 1, "AccoG": 1},
}

# Couleurs associées
PROFILE_COLORS: Dict[str, str] = {
    "2_voies": "#1f77b4",
    "2_voies_bretelle": "#ff7f0e",
    "3_voies": "#2ca02c",
    "3_voies_bretelle": "#9467bd",
    "4_voies": "#d62728",
    "4_voies_bretelle": "#8c564b",
    "Accès": "#17becf",
}

# 🧭 UX: profils PERSONNALISÉS créés par l'utilisateur — persistés en session et fusionnés
# aux profils de base à chaque exécution (Streamlit ré-exécute le script et réinitialise
# PROFILES/PROFILE_COLORS, donc cette fusion est rejouée proprement à chaque fois).
_BUILTIN_PROFILE_NAMES = set(PROFILES.keys())
st.session_state.setdefault("custom_profiles", {})        # {nom: {élément: comptage}}
st.session_state.setdefault("custom_profile_colors", {})  # {nom: "#hex"}
for _pname, _pdef in st.session_state["custom_profiles"].items():
    PROFILES[_pname] = _pdef
for _pname, _pcol in st.session_state["custom_profile_colors"].items():
    PROFILE_COLORS[_pname] = _pcol


# ─────────────────────────────────────────────────────────────
# Styles points PR (Fait/Ausculte) + légende
# ─────────────────────────────────────────────────────────────
PR_STYLE = {
    ('oui', 'oui'): dict(stroke='#27ae60', fill='#FFD700',
                         label='Fait = Oui, Ausculté = Oui'),
    ('oui', 'non'): dict(stroke='#27ae60', fill='#FFD700',
                         label='Fait = Oui, Ausculté = Non'),
    ('non', 'oui'): dict(stroke='#95a5a6', fill='#0000FF',
                         label='Fait = Non, Ausculté = Oui'),
    ('non', 'non'): dict(stroke='#95a5a6', fill='#FFFFFF',
                         label='Fait = Non, Ausculté = Non'),
}


def make_pr_points_legend(df: pd.DataFrame) -> str:
    """
    Construit une légende auto des points PR présents dans df (déjà filtré par Fait=Oui/Non).
    Affiche les 4 combinaisons possibles Fait/Ausculte, en n'affichant que celles présentes,
    avec le nombre de points par catégorie.
    """
    if df is None or df.empty:
        body = '<div style="color:#999;">Aucun point à afficher</div>'
    else:
        # comptages par (fait, ausc)
        tmp = df.copy()
        tmp["Fait"] = tmp["Fait"].map(_normalize_yn)
        tmp["Ausculte"] = tmp["Ausculte"].map(_normalize_yn)
        counts = tmp.groupby(["Fait", "Ausculte"]).size().to_dict()

        rows = []
        for key, sty in PR_STYLE.items():
            n = counts.get(key, 0)
            if n == 0:
                continue
            border = sty["stroke"]
            fill = sty["fill"]
            label = sty["label"]
            rows.append(f"""
              <div style="display:flex;align-items:center;margin:2px 0;">
                <svg width="18" height="18" style="margin-right:8px;">
                  <circle cx="9" cy="9" r="6" stroke="{border}" stroke-width="2" fill="{fill}" />
                </svg>
                <div>{label} <span style="color:#666">({n})</span></div>
              </div>
            """)

        if rows:
            body = "".join(rows)
        else:
            body = '<div style="color:#999;">Aucun point après filtrage</div>'

    # Afficher la note "anneau rouge" seulement si des points A_refaire=Oui sont présents
    _has_arefaire = (
        df is not None and not df.empty
        and "A_refaire" in df.columns
        and (df["A_refaire"].astype(str).str.strip().str.lower() == "oui").any()
    )
    _arefaire_note = ", Anneau rouge = À refaire" if _has_arefaire else ""

    # On place la légende en bas-gauche (celle des profils est déjà en bas-droite)
    html = f"""
    <div id="pr-maplegend" class="maplegend"
         style="position:absolute; z-index:9999; left:20px; bottom:100px;
                border:2px solid #bbb; background-color:rgba(255,255,255,0.9);
                border-radius:6px; padding:10px; font-size:15px; max-width:280px;">
      <div style="font-weight:700; margin-bottom:6px;">PR – Statuts</div>
      {body}
      <div style="margin-top:6px;color:#666;font-size:11px;">
        Bordure = Fait, Remplissage = Ausculté{_arefaire_note}
      </div>
    </div>
    """
    return html


def _normalize_yn(v: str) -> str:
    return (str(v or '').strip().lower()
            .replace('oui', 'oui')
            .replace('non', 'non'))


# =========================
# Matériaux par défaut (densités éditables) pour reprofilage
# =========================
DEFAULT_MATERIALS = [
    # Densités usuelles (ajustables) en t/m³ — prix unitaire en €/t (paramétrable, 0 par défaut)
    {"matériau": "GB",   "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # Grave-bitume
    {"matériau": "BBTM", "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # Très mince
    {"matériau": "BBM",  "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # Béton bitumineux mince (optionnel)
    {"matériau": "BBSG", "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # BBSG/BBSGF … (optionnel)
    {"matériau": "BBDr",   "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # Béton Bitumineux Drainant (plus léger)
    {"matériau": "BBME",   "densité_t_m3": 2.35, "épaisseur_cm": 0.0, "prix_eur_t": 0.0},   # Béton Bitumineux à Module Elevé
    {"matériau": "Grille anti-fissure", "densité_t_m3": 0.0, "épaisseur_cm": 0.01, "prix_eur_t": 0.0}, # Géogrille (masse surfacique)
]

# État (session) pour la table matériaux
if "materials_df" not in st.session_state:
    st.session_state["materials_df"] = pd.DataFrame(DEFAULT_MATERIALS)



# Transformers Lambert 93 <-> WGS84
TO_WGS84 = Transformer.from_crs(2154, 4326, always_xy=True)
TO_L93 = Transformer.from_crs(4326, 2154, always_xy=True)
# ─────────────────────────────────────────────────────────────
# Couche DIRMED — outils & villes (coordonnées WGS84, même système que la carte Folium)
# ─────────────────────────────────────────────────────────────
BOUCHES_DU_RHONE_CITIES = {
    "Marseille": (43.2965, 5.3698),
    "Aix-en-Provence": (43.5297, 5.4474),
    "Arles": (43.6766, 4.6278),
    "Martigues": (43.4058, 5.0480),
    "Salon-de-Provence": (43.6400, 5.0970),
    "Istres": (43.5167, 4.9833),
    "Vitrolles": (43.4600, 5.2489),
    "Miramas": (43.5833, 5.0000),
    "Fos-sur-Mer": (43.5333, 4.9333),
    "Chateauneuf-les-martigues": (43.38835, 5.1492),
    "Aubagne": (43.293046, 5.56842),
    "Paris": (48.8566, 2.3522),
    "Lyon": (45.7640, 4.8357),
    "Toulouse": (43.6045, 1.4440),
    "Nice": (43.7102, 7.2620),
    "Nantes": (47.2184, -1.5536),
    "Montpellier": (43.6119, 3.8777),
    "Strasbourg": (48.5734, 7.7521),
    "Bordeaux": (44.8378, -0.5792),
    "Lille": (50.6292, 3.0573),
    "Rennes": (48.1173, -1.6778),
    "Reims": (49.2583, 4.0317),
    "Le Havre": (49.4944, 0.1079),
    "Saint-Etienne": (45.4397, 4.3872),
    "Toulon": (43.1242, 5.9280),
}

@st.cache_data(show_spinner=False)  # ⚡ FLUIDITÉ: ne reprojette les ~23k points qu'au changement de données/filtres
def _prep_dirmed_df(df_src: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Prépare un DataFrame pour la couche DIRMED à partir du df déjà chargé :
    - garde colonnes utiles (avec colonnes optionnelles si présentes),
    - convertit x,y (virgule -> point) en float,
    - projette en WGS84 via TO_WGS84 (EPSG:2154 -> EPSG:4326, always_xy=True),
    - normalise Fait/Ausculte/A_refaire en minuscule.
    Retourne None si colonnes minimales absentes.
    """
    needed = {"route", "pr", "x", "y", "cote", "Fait", "Ausculte"}
    if not needed.issubset(set(df_src.columns)):
        return None

    optional = ["structure", "A_refaire", "V_L", "V_M", "V_R"]
    cols = list(needed) + [c for c in optional if c in df_src.columns]
    d = df_src[cols].copy()

    # Convertir x,y
    for c in ["x", "y"]:
        d[c] = (
            d[c].astype(str)
                .str.replace("\u00a0", "", regex=False)
                .str.replace(",", ".", regex=False)
                .str.strip()
        )
    d["x"] = pd.to_numeric(d["x"], errors="coerce")
    d["y"] = pd.to_numeric(d["y"], errors="coerce")

    # Nettoyage affichage
    d["route"] = d["route"].astype(str).str.strip()
    d["pr"] = d["pr"].astype(str).str.replace(",", ".", regex=False).str.strip()
    d["cote"] = d["cote"].astype(str).str.strip()

    # Normalisation Oui/Non
    for c in ["Fait", "Ausculte", "A_refaire"]:
        if c in d.columns:
            d[c] = d[c].astype(str).str.strip().str.lower()

    d = d.dropna(subset=["x", "y"])
    if d.empty:
        return None

    # Lambert-93 -> WGS84
    lons, lats = TO_WGS84.transform(d["x"].to_numpy(), d["y"].to_numpy())
    d["lat"] = lats
    d["lon"] = lons
    return d

# =========================
# Utilitaires
# =========================
def planimetric_distance_l93(coords_l93: List[Tuple[float, float]]) -> float:
    """Somme des distances euclidiennes entre points successifs (x,y) en Lambert93."""
    if len(coords_l93) < 2:
        return 0.0
    return float(
        sum(
            hypot(coords_l93[i + 1][0] - coords_l93[i][0],
                  coords_l93[i + 1][1] - coords_l93[i][1])
            for i in range(len(coords_l93) - 1)
        )
    )




def build_pr_popup_html(r: pd.Series) -> str:
    route = str(r.get("route", "") or "").strip()
    pr = str(r.get("pr", "") or "").strip()
    cote = str(r.get("cote", "") or "").strip()
    struct = str(r.get("structure") or "Non renseignée").strip()

    fait = str(r.get("Fait", "") or "").strip().lower()
    ausc = str(r.get("Ausculte", "") or "").strip().lower()
    refaire = str(r.get("A_refaire", "") or "").strip().lower()

    v_l = str(r.get("V_L", "") or "").strip()
    v_m = str(r.get("V_M", "") or "").strip()
    v_r = str(r.get("V_R", "") or "").strip()

    # 🧭 UX: lien d'accès Google Maps vers la position exacte du PR (depuis lat/lon)
    try:
        _lat = float(r.get("lat")); _lon = float(r.get("lon"))
        maps_url = f"https://www.google.com/maps/search/?api=1&query={_lat:.6f},{_lon:.6f}"
    except Exception:
        maps_url = ""
    maps_link_html = (
        f'<div style="margin-top:8px;"><a href="{maps_url}" target="_blank" rel="noopener" '
        f'style="display:inline-block;background:#1a73e8;color:#fff;padding:5px 10px;'
        f'border-radius:6px;font-size:12px;text-decoration:none;font-weight:600;">'
        f'📍 Ouvrir dans Google Maps</a></div>'
    ) if maps_url else ""

    def badge(label: str, ok: bool) -> str:
        color_bg = "#eafaf1" if ok else "#fdecea"
        color_fg = "#1e824c" if ok else "#c0392b"
        icon = "✔️" if ok else "❌"
        return (
            f'<span style="display:inline-block;padding:3px 8px;border-radius:12px;'
            f'font-weight:600;font-size:12px;background:{color_bg};color:{color_fg};'
            f'border:1px solid {color_fg}22;margin-right:6px;">{icon} {label}</span>'
        )

    badge_ausc = badge("Ausculte", ausc == "oui")
    badge_fait = badge("Fait", fait == "oui")
    badge_refaire = badge("À refaire", refaire == "oui")

    html = f"""
    <div style="
        font-family: Arial, sans-serif;
        font-size: 13px; color: #2c3e50; line-height: 1.35;
        background: #ffffff; border-radius: 10px; padding: 10px;
        box-shadow: 0 6px 20px rgba(0,0,0,0.15); min-width: 160px; max-width: 300px;">
      
      <div style="font-weight: 800; font-size: 10px; margin-bottom: 6px; color:#1f2d3d;">
        PR {pr} — {route} <span style="color:#7f8c8d;">({cote})</span>
      </div>

      <!-- Page 1 -->
      <div id="page1">
        <table style="width:100%; border-collapse: collapse; margin-bottom: 8px;">
          <tr>
            <td style="padding:4px 8px;color:#566573;"><b>Structure</b></td>
            <td style="padding:4px 8px;color:#2c3e50;">{struct}</td>
          </tr>
        </table>
        <div>{badge_ausc}{badge_fait}{badge_refaire}</div>
        {maps_link_html}
        <div style="text-align:right;margin-top:8px;">
          <a href="#" onclick="document.getElementById('page1').style.display='none';
                               document.getElementById('page2').style.display='block';return false;"
             style="color:#2980b9;font-size:12px;text-decoration:none;">➡ Structure conseillée</a>
        </div>
      </div>

      <!-- Page 2 -->
      <div id="page2" style="display:none;">
        <div style="font-weight:bold;margin-bottom:6px;">Structure conseillée</div>
        <table style="width:100%; border-collapse: collapse; margin-bottom: 8px;">
          <tr>
            <td style="padding:4px 8px;color:#566573;"><b>V_L</b></td>
            <td style="padding:4px 8px;color:#2c3e50;">{v_l or "—"}</td>
          </tr>
          <tr>
            <td style="padding:4px 8px;color:#566573;"><b>V_M</b></td>
            <td style="padding:4px 8px;color:#2c3e50;">{v_m or "—"}</td>
          </tr>
          <tr>
            <td style="padding:4px 8px;color:#566573;"><b>V_R</b></td>
            <td style="padding:4px 8px;color:#2c3e50;">{v_r or "—"}</td>
          </tr>
        </table>
        <div style="text-align:right;">
          <a href="#" onclick="document.getElementById('page2').style.display='none';
                               document.getElementById('page1').style.display='block';return false;"
             style="color:#2980b9;font-size:12px;text-decoration:none;">⬅ Retour</a>
        </div>
      </div>
    </div>
    """
    return html






# >>> MODIF : helper pour convertir l'écart de PR en mètres
def pr_delta_m(pr_start: float, pr_end: float) -> float:
    """Retourne 1000 * (PR_fin - PR_début)."""
    try:
        return 1000.0 * (float(pr_end) - float(pr_start))
    except Exception:
        return 0.0

@st.cache_data(show_spinner=False)
def load_pr_file(uploaded_file) -> pd.DataFrame:
    """
    Charge un fichier PR CSV (sep=';', decimal=',') ou Excel.
    Normalise: route, cote, pr, x, y, chainage_m (optionnelle).
    - 'cumul' est mappé vers 'chainage_m' si présent.
    """
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, sep=";", decimal=",")
    else:
        xls = pd.ExcelFile(uploaded_file)
        sheet_pr = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_pr)

    df = df.rename(
        columns={
            "Route": "route", "ROUTE": "route",
            "Cote": "cote", "COTE": "cote",
            "PR": "pr", "Pr": "pr",
            "X": "x", "Y": "y",
            "Chainage_m": "chainage_m", "CHAINAGE_M": "chainage_m", "chainage_m": "chainage_m",
            "cumul": "chainage_m",
        }
    )
    required = ["route", "cote", "pr", "x", "y"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes: {missing}. Requis: {required} (+ chainage_m optionnelle).")

    df["route"] = df["route"].astype(str)
    df["cote"] = df["cote"].astype(str)
    df["pr"] = pd.to_numeric(df["pr"], errors="coerce")
    df["x"] = pd.to_numeric(df["x"], errors="coerce")
    df["y"] = pd.to_numeric(df["y"], errors="coerce")
    if "chainage_m" in df.columns:
        df["chainage_m"] = pd.to_numeric(df["chainage_m"], errors="coerce")

    df = df.dropna(subset=["route", "cote", "pr", "x", "y"])
    df = df.sort_values(["route", "cote", "pr"]).reset_index(drop=True)
    # Dtypes plus compacts (quand présents)
    for col in ["route", "cote"]:
        if col in df.columns:
            df[col] = df[col].astype("category")
    for col in ["Gestionnaire", "departement"]:
        if col in df.columns:
            df[col] = df[col].astype("category")

    return df

@st.cache_data(show_spinner=False)
def load_overrides_file(uploaded_file) -> Optional[pd.DataFrame]:
    """
    Charge un fichier d'overrides (CSV/Excel).
    Colonnes: route, cote, pr_start, pr_end, element, largeur_m
    Si Excel et une feuille 'overrides' existe (casse ignorée), on la prend.
    """
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        ov = pd.read_csv(uploaded_file, sep=";", decimal=",")
    else:
        xls = pd.ExcelFile(uploaded_file)
        lower_names = [s.lower() for s in xls.sheet_names]
        if "overrides" in lower_names:
            idx = lower_names.index("overrides")
            sheet_name = xls.sheet_names[idx]
        else:
            sheet_name = xls.sheet_names[0]
        ov = pd.read_excel(xls, sheet_name=sheet_name)

    ov = ov.rename(
        columns={
            "Route": "route", "Cote": "cote",
            "PR_start": "pr_start", "Pr_start": "pr_start",
            "PR_end": "pr_end", "Pr_end": "pr_end",
            "Element": "element", "element": "element",
            "Largeur_m": "largeur_m", "largeur_m": "largeur_m",
        }
    )
    needed_ov = ["route", "cote", "pr_start", "pr_end", "element", "largeur_m"]
    missing_ov = [c for c in needed_ov if c not in ov.columns]
    if missing_ov:
        raise ValueError(f"Overrides: colonnes manquantes {missing_ov}. Attendu: {needed_ov}")

    ov["route"] = ov["route"].astype(str)
    ov["cote"] = ov["cote"].astype(str)
    ov["pr_start"] = pd.to_numeric(ov["pr_start"], errors="coerce")
    ov["pr_end"] = pd.to_numeric(ov["pr_end"], errors="coerce")
    ov["element"] = ov["element"].astype(str)
    ov["largeur_m"] = pd.to_numeric(ov["largeur_m"], errors="coerce")
    ov = ov.dropna()
    return ov

def midpoint_wgs(coords_wgs: List[Tuple[float, float]]) -> Tuple[float, float]:
    lats = [c[0] for c in coords_wgs]
    lons = [c[1] for c in coords_wgs]
    return float(np.mean(lats)), float(np.mean(lons))

def l93_to_wgs(latlon_l93: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    # input (x,y) -> output (lat, lon)
    return [TO_WGS84.transform(x, y)[::-1] for x, y in latlon_l93]

def wgs_to_l93(coords_wgs: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    # input (lat, lon) -> output (x,y)
    return [TO_L93.transform(lon, lat) for lat, lon in coords_wgs]

def merge_profile_mix(profiles_selected: List[str], percentages: List[float]) -> Dict[str, float]:
    weights = np.array(percentages, dtype=float)
    if weights.sum() == 0:
        return {}
    weights = weights / weights.sum()
    agg: Dict[str, float] = {}
    for prof_name, w in zip(profiles_selected, weights):
        prof = PROFILES[prof_name]
        for elem, count in prof.items():
            agg[elem] = agg.get(elem, 0.0) + float(w) * float(count)
    return agg

def dominant_profile_name(profiles_selected: List[str], percentages: List[float]) -> Optional[str]:
    """Profil dominant (max %) ; si égalité ou vide -> None."""
    if not profiles_selected:
        return None
    if len(percentages) != len(profiles_selected):
        return profiles_selected[0]
    if sum(percentages) == 0:
        return None
    max_idx = int(np.argmax(percentages))
    return profiles_selected[max_idx]

def apply_overrides(
    widths: Dict[str, float],
    overrides: Optional[pd.DataFrame],
    route: str,
    cote: str,
    pr_start: float,
    pr_end: float,
) -> Dict[str, float]:
    if overrides is None:
        return widths
    mask = (
        (overrides["route"] == route)
        & (overrides["cote"] == cote)
        & (overrides["pr_start"] <= pr_end)
        & (overrides["pr_end"] >= pr_start)
    )
    ov = overrides.loc[mask]
    if ov.empty:
        return widths
    new_widths = widths.copy()
    for _, row in ov.iterrows():
        elem = str(row["element"])
        val = float(row["largeur_m"])
        if elem in new_widths:
            new_widths[elem] = val
    return new_widths

def compute_areas(
    distance_m: float,
    widths_m: Dict[str, float],
    element_counts: Dict[str, float],
    included_elements: List[str],
) -> Tuple[pd.DataFrame, float]:
    rows = []
    total = 0.0
    for elem, count in element_counts.items():
        if elem not in included_elements:
            continue
        width = float(widths_m.get(elem, 0.0))
        width_equiv = float(count) * width
        area = float(distance_m) * width_equiv
        rows.append(
            {
                "element": elem,
                "count_equiv": round(float(count), 3),
                "width_m": round(width, 3),
                "width_equiv_m": round(width_equiv, 3),
                "area_m2": area,
            }
        )
        total += area
    df = pd.DataFrame(rows).sort_values("element").reset_index(drop=True)
    return df, float(total)

def build_segment_key(route: str, cote: str, pr_start: float, pr_end: float) -> str:
    return f"{route}__{cote}__{pr_start}->({pr_end})"

def _poly_hash(poly, precision: int = 6) -> str:
    return '|'.join(f'{round(lat, precision)},{round(lon, precision)}' for lat, lon in poly)

def _dedup_polylines(polys, precision: int = 6):
    seen = set(); out = []
    for p in polys:
        if not p or len(p) < 2:
            continue
        h = _poly_hash(p, precision)
        if h not in seen:
            seen.add(h); out.append(p)
    return out

def parse_drawn_polylines(map_data: Optional[Dict[str, Any]], prefer_last: bool = True) -> List[List[Tuple[float, float]]]:
    if not map_data:
        return []
    polys: List[List[Tuple[float, float]]]=[]

    def _add_from_geom(geom: Dict[str, Any]):
        gtype = (geom or {}).get("type")
        coords = (geom or {}).get("coordinates", [])
        if gtype == "LineString" and coords:
            pts = [(lat, lon) for lon, lat in coords]
            if len(pts) >= 2:
                polys.append(pts)
        elif gtype == "MultiLineString":
            for line in coords or []:
                pts = [(lat, lon) for lon, lat in line]
                if len(pts) >= 2:
                    polys.append(pts)
        elif gtype == "FeatureCollection":
            for f in (geom.get("features") or []):
                _add_from_geom((f or {}).get("geometry") or {})

    last = map_data.get("last_active_drawing") or {}
    drawings = map_data.get("all_drawings") or []

    if prefer_last and last.get("geometry"):
        _add_from_geom(last["geometry"])
    for feat in drawings:
        _add_from_geom((feat or {}).get("geometry") or {})

    return _dedup_polylines(polys)

def parse_drawn_polygons(map_data: Optional[Dict[str, Any]]) -> List[List[Tuple[float, float]]]:
    """
    Extrait les polygones Leaflet.Draw depuis st_folium (Polygon / MultiPolygon).
    Retourne une liste d’anneaux (lat,lon) pour les contours extérieurs.
    """
    if not map_data:
        return []
    polys: List[List[Tuple[float, float]]] = []

    def _add_from_geom(geom: Dict[str, Any]):
        if not geom:
            return
        gtype = geom.get("type")
        coords = geom.get("coordinates", [])
        if gtype == "Polygon" and coords:
            ring = coords[0]
            pts = [(lat, lon) for lon, lat in ring]
            if len(pts) >= 3:
                polys.append(pts)
        elif gtype == "MultiPolygon":
            for poly in coords:
                if poly:
                    ring = poly[0]
                    pts = [(lat, lon) for lon, lat in ring]
                    if len(pts) >= 3:
                        polys.append(pts)
        elif gtype == "FeatureCollection":
            for f in (geom.get("features") or []):
                _add_from_geom((f or {}).get("geometry") or {})

    last = (map_data or {}).get("last_active_drawing") or {}
    drawings = (map_data or {}).get("all_drawings") or []
    if last.get("geometry"):
        _add_from_geom(last["geometry"])
    for feat in drawings:
        _add_from_geom((feat or {}).get("geometry") or {})

    # Déduplication (arrondi)
    def _h(poly, prec=6):
        return "\n".join(f"{round(lat,prec)},{round(lon,prec)}" for lat,lon in poly)
    out, seen = [], set()
    for p in polys:
        if len(p) < 3:
            continue
        h = _h(p)
        if h not in seen:
            seen.add(h); out.append(p)
    return out

def polygon_area_m2_from_wgs(ring_wgs: List[Tuple[float, float]] , TO_L93=TO_L93) -> float:
    """
    Aire planimétrique (m²) d’un polygone (lat,lon) en projetant en Lambert‑93 (shoelace).
    """
    if not ring_wgs or len(ring_wgs) < 3:
        return 0.0
    xs, ys = [], []
    for lat, lon in ring_wgs:
        x, y = TO_L93.transform(lon, lat)
        xs.append(float(x)); ys.append(float(y))
    # fermer l’anneau si besoin
    if xs[0] != xs[-1] or ys[0] != ys[-1]:
        xs.append(xs[0]); ys.append(ys[0])
    area = 0.0
    for i in range(len(xs)-1):
        area += xs[i]*ys[i+1] - xs[i+1]*ys[i]
    return abs(area) * 0.5

def make_legend_html(selected: List[str], percentages: List[int], show_percentages: bool = True) -> str:
    rows = []
    sel_pct = {name: pct for name, pct in zip(selected, percentages)}
    for key, color in PROFILE_COLORS.items():
        label = key.replace("_", " ")
        pct_display = f" — {sel_pct.get(key, 0)}%" if (key in sel_pct and show_percentages) else ""
        weight = "font-weight:600;" if key in sel_pct else "font-weight:400;"
        rows.append(
            f"""
<div style="display:flex;align-items:center;margin:2px 0;{weight}">
  <div style="width:14px;height:14px;background:{color};
  border:1px solid #333;margin-right:8px;"></div>
  <div>{label}{pct_display}</div>
</div>
"""
        )
    return f"""
<div id="maplegend" class="maplegend"
 style="position:absolute;z-index:9999;border:2px solid #bbb;
 background-color:rgba(255,255,255,0.9);border-radius:6px;padding:10px;font-size:12px;right:20px;bottom:20px;">
  <div class="legend-title" style="font-weight:700;margin-bottom:6px;">Profils & couleurs</div>
  <div class="legend-scale">{''.join(rows)}</div>
  <div style="margin-top:6px;color:#666;">*En gras : profils sélectionnés.</div>
</div>
"""


# =========================
# UI — Flux unique
# =========================
st.title("PR-ESTIME")

# =========================
# 🧭 UX (F1): Sauvegarde / reprise de projet — sérialise l'état métier en JSON.
# Additif : n'altère aucun flux existant, ajoute seulement export/import.
# =========================
_PROJECT_STATE_KEYS = [
    "subsegments", "surfaces", "edited_geoms", "edited_polygons",
    "profile_counts", "circles", "profiles_selected", "percents",
    "included_elements", "widths_applied", "dom_name", "seg_color",
    "custom_profiles", "custom_profile_colors",  # 🧭 UX: profils personnalisés/modifiés
]
_PROJECT_KEY_PREFIXES = ("rabot_list_", "mats_", "surf_mats_")


def _project_snapshot() -> dict:
    """Capture l'état métier courant (sous-segments, surfaces, passes, matériaux…)."""
    sst = st.session_state
    snap = {"_version": 1, "keys": {}, "materials_df": None}
    for k in _PROJECT_STATE_KEYS:
        if k in sst:
            snap["keys"][k] = sst[k]
    for k in list(sst.keys()):
        if isinstance(k, str) and k.startswith(_PROJECT_KEY_PREFIXES):
            snap["keys"][k] = sst[k]
    mdf = sst.get("materials_df")
    if isinstance(mdf, pd.DataFrame):
        snap["materials_df"] = mdf.to_dict(orient="list")
    return snap


def _project_to_json(snap: dict) -> str:
    import json

    def _default(o):
        if isinstance(o, (np.integer,)):
            return int(o)
        if isinstance(o, (np.floating,)):
            return float(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        if isinstance(o, tuple):
            return list(o)
        return str(o)

    return json.dumps(snap, ensure_ascii=False, indent=2, default=_default)


def _project_restore(snap: dict) -> int:
    """Réinjecte un snapshot dans st.session_state. Retourne le nb d'éléments restaurés."""
    n = 0
    for k, v in (snap or {}).get("keys", {}).items():
        st.session_state[k] = v
        n += 1
    mdf = (snap or {}).get("materials_df")
    if mdf:
        st.session_state["materials_df"] = pd.DataFrame(mdf)
        n += 1
    return n


# 🧭 UX: Générateur de RAPPORT EXCEL propre à 4 onglets (Surface général, Rabotage, Reprofilage, Surface)
def _build_report_xlsx(context: dict,
                       df_detail: pd.DataFrame,
                       recap_elements: pd.DataFrame,
                       df_rabot: pd.DataFrame,
                       df_reprof: pd.DataFrame,
                       df_surface: pd.DataFrame,
                       df_cost: pd.DataFrame = None) -> bytes:
    """
    Rapport Excel mis en forme à 4 onglets :
      1) Surface général : contexte + total voirie + surfaces par élément / profil / sous-segment
      2) Rabotage        : détail des passes + totaux par élément et par hauteur
      3) Reprofilage     : détail par matériau + totaux par élément et par matériau
      4) Surface         : surfaces dessinées (polygones) + totaux
    Lecture simple : titres, en-têtes colorés, totaux en évidence.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Charte VINCI Construction : bleu marine #004489, rouge #E20025, police Arial
    NAVY = "004489"; RED = "E20025"; SECTION = "E6ECF4"; TOTAL = "FBE2E6"
    f_title = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    f_head = Font(name="Arial", bold=True, color="FFFFFF")
    f_section = Font(name="Arial", bold=True, size=11, color="004489")
    f_total = Font(name="Arial", bold=True, color="004489")
    f_kpi = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    f_cell = Font(name="Arial")
    fill_navy = PatternFill("solid", fgColor=NAVY)
    fill_section = PatternFill("solid", fgColor=SECTION)
    fill_total = PatternFill("solid", fgColor=TOTAL)
    fill_kpi = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    FMT_M2 = '#,##0'
    FMT_DEC = '#,##0.00'
    FMT_BY_COL = {
        "surface_m2": FMT_M2, "aire_m2": FMT_M2, "largeur_équivalente_m": FMT_DEC,
        "largeur_m": FMT_DEC, "comptage_équivalent": FMT_DEC, "distance_m": '#,##0',
        "volume_m3": FMT_M2, "vol_rabot_m3": FMT_M2, "vol_reprof_m3": FMT_M2,
        "tonnage_t": FMT_M2, "hauteur_cm": FMT_DEC, "épaisseur_cm": FMT_DEC,
        "densité_t_m3": FMT_DEC, "rabot_h_cm": FMT_DEC,
        "surface_cumulée_m2": FMT_M2, "volume_cumulé_m3": FMT_M2,
        "prix_eur_t": FMT_DEC, "coût_eur": FMT_M2,
    }

    def _fmt_for(col):
        return FMT_BY_COL.get(str(col), None)

    wb = Workbook()

    def _title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
        c = ws.cell(row=1, column=1, value=text)
        c.font = f_title; c.fill = fill_navy; c.alignment = center
        ws.row_dimensions[1].height = 26

    def _section(ws, row, text, ncols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
        c = ws.cell(row=row, column=1, value=text)
        c.font = f_section; c.fill = fill_section; c.alignment = left
        return row + 1

    def _kpi(ws, row, label, value, fmt=FMT_M2):
        a = ws.cell(row=row, column=1, value=label); a.font = f_kpi; a.fill = fill_kpi
        b = ws.cell(row=row, column=2, value=value); b.font = f_kpi; b.fill = fill_kpi
        if fmt and isinstance(value, (int, float)):
            b.number_format = fmt
        return row + 1

    def _table(ws, row, df, total_cols=None, total_label="TOTAL"):
        if df is None or len(df) == 0:
            ws.cell(row=row, column=1, value="(aucune donnée)").font = Font(italic=True, color="888888")
            return row + 2
        cols = list(df.columns)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=row, column=j, value=str(col))
            c.font = f_head; c.fill = fill_navy; c.alignment = center; c.border = border
        r = row + 1
        for _, srow in df.iterrows():
            for j, col in enumerate(cols, start=1):
                val = srow[col]
                try:
                    import numpy as _np
                    if isinstance(val, (_np.integer,)):
                        val = int(val)
                    elif isinstance(val, (_np.floating,)):
                        val = float(val)
                except Exception:
                    pass
                if not isinstance(val, (int, float, str)) or val is None:
                    val = "" if val is None else str(val)
                c = ws.cell(row=r, column=j, value=val)
                c.border = border
                c.font = f_cell
                fmt = _fmt_for(col)
                if fmt and isinstance(val, (int, float)):
                    c.number_format = fmt
            r += 1
        if total_cols:
            for j, col in enumerate(cols, start=1):
                c = ws.cell(row=r, column=j)
                c.fill = fill_total; c.font = f_total; c.border = border
                if j == 1:
                    c.value = total_label
                elif col in total_cols:
                    try:
                        c.value = float(pd.to_numeric(df[col], errors="coerce").sum())
                        fmt = _fmt_for(col)
                        if fmt:
                            c.number_format = fmt
                    except Exception:
                        pass
            r += 1
        for j, col in enumerate(cols, start=1):
            maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
            ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen + 2, 10), 26)
        return r + 1

    # ---- Onglet 1 : Surface général ----
    ws1 = wb.active
    ws1.title = "Surface général"
    _title(ws1, "RAPPORT D'ESTIMATION — SURFACES & MATÉRIAUX", 4)
    row = 3
    for label, val in [
        ("Route", context.get("route", "")),
        ("Côté", context.get("cote", "")),
        ("PR début → fin", f"{context.get('pr_start','')} → {context.get('pr_end','')}"),
        ("Distance (m)", context.get("distance_m", 0)),
        ("Méthode de distance", context.get("methode", "")),
        ("Profil dominant", context.get("profil_dominant", "")),
    ]:
        a = ws1.cell(row=row, column=1, value=label); a.font = f_total
        b = ws1.cell(row=row, column=2, value=val)
        if label == "Distance (m)" and isinstance(val, (int, float)):
            b.number_format = '#,##0'
        row += 1
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 26
    row += 1
    row = _kpi(ws1, row, "SURFACE TOTALE VOIRIE (m²)", float(context.get("surface_totale_voirie", 0.0)))
    row += 1
    row = _section(ws1, row, "Surface par élément (voie)", 2)
    re_df = recap_elements.copy() if recap_elements is not None else pd.DataFrame(columns=["élément", "surface_m2"])
    row = _table(ws1, row, re_df, total_cols=["surface_m2"])
    if df_detail is not None and not df_detail.empty and "profil_nom" in df_detail.columns:
        row = _section(ws1, row, "Surface par profil", 2)
        by_prof = (df_detail.groupby("profil_nom", as_index=False)["surface_m2"].sum()
                   .sort_values("surface_m2", ascending=False))
        row = _table(ws1, row, by_prof, total_cols=["surface_m2"])
    if df_detail is not None and not df_detail.empty and "sous_segment" in df_detail.columns:
        row = _section(ws1, row, "Surface par sous-segment", 2)
        by_ss = (df_detail.groupby("sous_segment", as_index=False)["surface_m2"].sum()
                 .sort_values("sous_segment"))
        row = _table(ws1, row, by_ss, total_cols=["surface_m2"])
    if df_detail is not None and not df_detail.empty:
        row = _section(ws1, row, "Détail complet", len(df_detail.columns))
        row = _table(ws1, row, df_detail, total_cols=["surface_m2"])
    ws1.freeze_panes = "A3"

    # ---- Onglet 2 : Rabotage ----
    ws2 = wb.create_sheet("Rabotage")
    _title(ws2, "RABOTAGE — Volumes par passe", 5)
    row = 3
    if df_rabot is not None and not df_rabot.empty:
        row = _kpi(ws2, row, "VOLUME TOTAL RABOTAGE (m³)",
                   float(pd.to_numeric(df_rabot["volume_m3"], errors="coerce").sum()), fmt=FMT_M2)
        row += 1
        row = _section(ws2, row, "Détail des passes", len(df_rabot.columns))
        row = _table(ws2, row, df_rabot, total_cols=["volume_m3"])
        row = _section(ws2, row, "Totaux par élément", 2)
        te = (df_rabot.groupby("élément", as_index=False)["volume_m3"].sum()
              .sort_values("volume_m3", ascending=False))
        row = _table(ws2, row, te, total_cols=["volume_m3"])
        if "hauteur_cm" in df_rabot.columns:
            row = _section(ws2, row, "Totaux par hauteur de rabotage", 3)
            th = (df_rabot.assign(hauteur_cm=pd.to_numeric(df_rabot["hauteur_cm"], errors="coerce").round(2))
                  .groupby("hauteur_cm", as_index=False)[["surface_m2", "volume_m3"]].sum()
                  .sort_values("hauteur_cm"))
            row = _table(ws2, row, th, total_cols=["surface_m2", "volume_m3"])
    else:
        ws2.cell(row=row, column=1, value="Aucune passe de rabotage saisie.").font = Font(italic=True, color="888888")
    ws2.freeze_panes = "A3"

    # ---- Onglet 3 : Reprofilage ----
    ws3 = wb.create_sheet("Reprofilage")
    _title(ws3, "REPROFILAGE — Volumes & tonnages par matériau", 7)
    row = 3
    if df_reprof is not None and not df_reprof.empty:
        row = _kpi(ws3, row, "VOLUME TOTAL (m³)",
                   float(pd.to_numeric(df_reprof["volume_m3"], errors="coerce").sum()), fmt=FMT_M2)
        row = _kpi(ws3, row, "TONNAGE TOTAL (t)",
                   float(pd.to_numeric(df_reprof["tonnage_t"], errors="coerce").sum()), fmt=FMT_M2)
        row += 1
        row = _section(ws3, row, "Détail par matériau", len(df_reprof.columns))
        row = _table(ws3, row, df_reprof, total_cols=["volume_m3", "tonnage_t"])
        row = _section(ws3, row, "Totaux par élément", 3)
        te = (df_reprof.groupby("élément", as_index=False)[["volume_m3", "tonnage_t"]].sum()
              .sort_values("tonnage_t", ascending=False))
        row = _table(ws3, row, te, total_cols=["volume_m3", "tonnage_t"])
        row = _section(ws3, row, "Totaux par matériau", 3)
        tm = (df_reprof.groupby("matériau", as_index=False)[["volume_m3", "tonnage_t"]].sum()
              .sort_values("tonnage_t", ascending=False))
        row = _table(ws3, row, tm, total_cols=["volume_m3", "tonnage_t"])
    else:
        ws3.cell(row=row, column=1, value="Aucun matériau/épaisseur saisi.").font = Font(italic=True, color="888888")
    ws3.freeze_panes = "A3"

    # ---- Onglet 4 : Surface ----
    ws4 = wb.create_sheet("Surface")
    _title(ws4, "SURFACES DESSINÉES (polygones)", 6)
    row = 3
    if df_surface is not None and not df_surface.empty:
        tcols = [c for c in ["aire_m2", "vol_rabot_m3", "vol_reprof_m3", "tonnage_t"] if c in df_surface.columns]
        row = _section(ws4, row, "Surfaces & quantités", len(df_surface.columns))
        row = _table(ws4, row, df_surface, total_cols=tcols)
    else:
        ws4.cell(row=row, column=1, value="Aucune surface dessinée ajoutée.").font = Font(italic=True, color="888888")
    ws4.freeze_panes = "A3"

    # ---- Onglet 5 : Coût (€) ----
    ws5 = wb.create_sheet("Coût (€)")
    _title(ws5, "ESTIMATION DU COÛT — tonnage × prix unitaire", 4)
    row = 3
    if df_cost is not None and not df_cost.empty and "coût_eur" in df_cost.columns:
        row = _kpi(ws5, row, "COÛT TOTAL ESTIMÉ (€)",
                   float(pd.to_numeric(df_cost["coût_eur"], errors="coerce").sum()), fmt=FMT_M2)
        if "tonnage_t" in df_cost.columns:
            row = _kpi(ws5, row, "TONNAGE TOTAL (t)",
                       float(pd.to_numeric(df_cost["tonnage_t"], errors="coerce").sum()), fmt=FMT_M2)
        row += 1
        row = _section(ws5, row, "Coût par matériau", len(df_cost.columns))
        _ctot = [c for c in ["tonnage_t", "coût_eur"] if c in df_cost.columns]
        row = _table(ws5, row, df_cost, total_cols=_ctot)
    else:
        ws5.cell(row=row, column=1,
                 value="Renseigne les prix (€/t) dans la bibliothèque et le Reprofilage pour estimer un coût.").font = Font(italic=True, color="888888")
    ws5.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- Import des données
with st.container():
    st.markdown("#### Import des données")
    uploaded = st.file_uploader(
        "Fichier PR (CSV ou Excel) : colonnes route, cote, pr, x, y ; optionnelle : chainage_m "
        "(dans ton CSV, 'cumul' est automatiquement mappé en chainage_m).",
        type=["xlsx", "csv"],
    )
    col_ov1, col_ov2 = st.columns([1, 2])
    with col_ov1:
        use_overrides = st.checkbox("Ajouter des overrides (optionnel)")
    with col_ov2:
        overrides_file = None
        if use_overrides:
            overrides_file = st.file_uploader(
                "Fichier overrides (CSV/Excel) : route, cote, pr_start, pr_end, element, largeur_m",
                type=["xlsx", "csv"],
                key="ov_file",
            )

if not uploaded:
    st.info("➕ Charge un fichier PR pour commencer.")
    st.stop()

# ---- Lecture des fichiers
try:
    df = load_pr_file(uploaded)
    overrides = None
    if overrides_file is not None:
        overrides = load_overrides_file(overrides_file)
    st.success(
        f"✅ PR chargés : {len(df)} lignes, {df['route'].nunique()} route(s), {df['cote'].nunique()} côté(s)."
    )
except Exception as e:
    st.error(f"Erreur lors du chargement: {e}")
    st.stop()

# 🧭 UX (F1): sauvegarder / reprendre le travail (sous-segments, surfaces, passes, matériaux…)
with st.expander("💾 Projet — sauvegarder / reprendre votre travail", expanded=False):
    _cpa, _cpb = st.columns(2)
    with _cpa:
        st.download_button(
            "💾 Exporter le projet (.json)",
            data=_project_to_json(_project_snapshot()).encode("utf-8"),
            file_name="projet_estimation_egpf.json",
            mime="application/json",
            help="Enregistre tout votre travail pour le reprendre plus tard (ou sur un autre poste).",
            use_container_width=True,
        )
    with _cpb:
        _proj_file = st.file_uploader("📂 Reprendre un projet (.json)", type=["json"], key="proj_import")
        if _proj_file is not None:
            _marker = f"{_proj_file.name}:{getattr(_proj_file, 'size', 0)}"
            if st.session_state.get("_proj_last_import") != _marker:
                try:
                    import json
                    _snap_in = json.loads(_proj_file.getvalue().decode("utf-8"))
                    _n = _project_restore(_snap_in)
                    st.session_state["_proj_last_import"] = _marker
                    st.success(f"Projet repris ({_n} éléments restaurés).")
                    st.rerun()
                except Exception as _e:
                    st.error(f"Fichier projet illisible : {_e}")
    st.caption("Astuce : le travail est sinon perdu si vous rechargez la page. Pensez à exporter régulièrement.")


# =========================
# Filtres globaux (Gestionnaire, depPr, Route) + Sélection du segment (sidebar)
# =========================
with st.sidebar:
    st.header("Filtres")

    # Normalisation souple de colonnes éventuelles
    _rename_map_soft = {}
    for alt in ["depPr", "DEPARTEMENT", "departement", "Dept", "dept"]:
        if alt in df.columns and "departement" not in df.columns:
            _rename_map_soft[alt] = "departement"
            break
    for alt in ["Gestionnaire", "Gestionnaire", "concession", "Concession"]:
        if alt in df.columns and "Gestionnaire" not in df.columns:
            _rename_map_soft[alt] = "Gestionnaire"
            break
    if _rename_map_soft:
        df = df.rename(columns=_rename_map_soft)

    # Filtres Gestionnaire & Département
    cons_sel = []
    if "Gestionnaire" in df.columns:
        cons_opts = sorted(pd.Series(df["Gestionnaire"].dropna().astype(str).unique()).tolist())
        cons_sel = st.multiselect("Gestionnaires", options=cons_opts, default=[])

    dep_sel = []
    if "departement" in df.columns:
        dep_opts = sorted(pd.Series(df["departement"].dropna().astype(str).unique()).tolist())
        dep_sel = st.multiselect("Département", options=dep_opts, default=[])

    _df_f = df.copy()
    if cons_sel and "Gestionnaire" in _df_f.columns:
        _df_f = _df_f[_df_f["Gestionnaire"].astype(str).isin(cons_sel)]
    if dep_sel and "departement" in _df_f.columns:
        _df_f = _df_f[_df_f["departement"].astype(str).isin(dep_sel)]

    # Filtre routes (multi) pour réduire la liste proposée ensuite
    route_opts = sorted(pd.Series(_df_f["route"].dropna().astype(str).unique()).tolist())
    route_filter = st.multiselect("Routes (filtre)", options=route_opts, default=[])
    if route_filter:
        _df_f = _df_f[_df_f["route"].astype(str).isin(route_filter)]

    st.caption(f"📉 Lignes après filtres : {len(_df_f):,}".replace(',', ' '))

    # --- Sélection du segment (déplacée à gauche)
    st.markdown("---")
    with st.expander("🧭 Sélection du segment", expanded=True):
        if _df_f.empty:
            st.info("Aucune donnée après filtres (Gestionnaire/Département/Routes).")
        else:
            route = st.selectbox("Route", sorted(_df_f["route"].astype(str).unique()))
            cotes_dispo = _df_f.loc[_df_f["route"] == route, "cote"].astype(str).unique().tolist()
            cote = st.selectbox("Côté", sorted(cotes_dispo))

            subset = _df_f[( _df_f["route"] == route) & (_df_f["cote"] == cote)].sort_values("pr").reset_index(drop=True)
            prs = sorted(subset["pr"].dropna().unique().tolist())
            # 🧭 UX: curseur de plage PR (début → fin) — plus rapide et "fin > début" garanti.
            # Repli sur les deux listes d'origine si moins de 2 PR disponibles.
            if len(prs) >= 2:
                pr_start, pr_end = st.select_slider(
                    "PR (début → fin)",
                    options=prs,
                    value=(prs[0], prs[1]),  # même tronçon par défaut qu'avant
                    help="Glisse les deux poignées pour choisir le PR de début et le PR de fin.",
                )
                if pr_end == pr_start:  # garde-fou : éviter début == fin
                    later = [p for p in prs if p > pr_start]
                    pr_end = later[0] if later else pr_end
            else:
                pr_start = st.selectbox("PR début", prs, index=0 if prs else None)
                prs_after = [p for p in prs if p > pr_start] if prs else []
                pr_end = st.selectbox("PR fin", prs_after, index=0 if prs_after else None)

# Appliquer les filtres au DF global pour la suite
df = _df_f

if df.empty:
    st.warning("Aucune donnée après filtres (Gestionnaire/depPr/route).")
    st.stop()

# Validation de la sélection + calculs de base (identiques à avant)
sel = subset[subset["pr"].isin([pr_start, pr_end])].sort_values("pr")
if len(sel) != 2:
    st.warning("Sélectionne deux PR valides (début < fin).")
    st.stop()

pr1 = sel.iloc[0]
pr2 = sel.iloc[1]

coords_l93 = [(float(pr1["x"]), float(pr1["y"])), (float(pr2["x"]), float(pr2["y"]))]
coords_wgs = l93_to_wgs(coords_l93)
seg_key = build_segment_key(route, cote, float(pr_start), float(pr_end))

# 🧭 UX: bandeau résultat collant en haut (distance & surface du DERNIER calcul de ce tronçon)
_cache_seg = st.session_state.get("calc_cache", {}).get(seg_key)
_b_dist, _b_surf = "—", "—"
if _cache_seg is not None:
    _adf = _cache_seg.get("areas_df_fr")
    if _adf is not None and not _adf.empty and "surface_m2" in _adf.columns:
        _b_surf = f"{float(_adf['surface_m2'].sum()):,.0f}".replace(",", " ") + " m²"
        if "distance_m" in _adf.columns and _adf["distance_m"].notna().any():
            _b_dist = f"{float(_adf['distance_m'].dropna().iloc[0]):,.0f}".replace(",", " ") + " m"
st.markdown(
    f"""<div style="position:sticky;top:0;z-index:1000;background:#004489;color:#fff;
    padding:8px 14px;border-radius:8px;margin:6px 0;font-family:Arial;font-weight:600;
    display:flex;flex-wrap:wrap;gap:22px;align-items:center;">
    <span>🛣️ {route} {cote} · PR {int(float(pr_start))}→{int(float(pr_end))}</span>
    <span>📏 Distance : {_b_dist}</span>
    <span>🟦 Surface : {_b_surf}</span>
    <span style="font-weight:400;opacity:.8;">(dernier calcul)</span></div>""",
    unsafe_allow_html=True,
)

# --- Distance & courbure (masqué par défaut, accessible via bouton)
# État par défaut en session
ss = st.session_state
ss.setdefault("dist_panel_open", False)
ss.setdefault("filter_panel_open", False)  # NOUVEAU
ss.setdefault("circle_panel_open", False)  # NOUVEAU
ss.setdefault("dist_method", "Segment édité")
ss.setdefault("curvature_factor", 1.00)
ss.setdefault("map_height", 750)
ss.setdefault("zoom_init", 14)
ss.setdefault("legend_panel_open", False)
ss.setdefault("circles_panel_open", False)
# État par défaut en session (ajouter ces 2 lignes)
ss.setdefault("quick_profile_open", {})    # {seg_key: bool}
ss.setdefault("quick_profile_choice", {})  # {seg_key: "3_voies" ...}
ss.setdefault("quick_profile_apply", {})  # {seg_key: bool}
ss.setdefault("profile_panel_open", False)  # nouveau panneau "Profils & éléments"
ss.setdefault("profile_ready", {})  # {seg_key: bool} — prêt à ajouter un sous-segment ?
# Valeur par défaut pour la distance fixe (utilisée au premier rendu du champ)
default_fixed = pr_delta_m(pr_start, pr_end) or 1000.0

# ─────────────────────────────────────────────────────────────
# Barre d'actions — 3 boutons en ligne + style compact
# ─────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    /* Groupe compact : réduit la hauteur et le padding des st.button */
    .compact-buttons .stButton > button {
        padding: 6px 10px;
        min-height: 0;
        line-height: 1.1;
        font-size: 13px;
        border-radius: 6px;
    }
    /* Espace horizontal réduit entre colonnes */
    .compact-buttons [data-testid="column"] { gap: 0.25rem; }
    </style>
    """,
    unsafe_allow_html=True
)

# 🧭 UX (F2): fil d'étapes — repère visuel du parcours (présentation seule, aucun impact logique)
st.markdown(
    '<div style="display:flex;flex-wrap:wrap;gap:10px;margin:4px 0 2px 0;font-family:Arial;font-size:13px;">'
    '<span style="background:#e7f0f7;border-radius:6px;padding:3px 9px;">1 · Importer ✓</span>'
    '<span style="background:#e7f0f7;border-radius:6px;padding:3px 9px;">2 · Choisir le tronçon (panneau de gauche)</span>'
    '<span style="background:#dceaf6;border-radius:6px;padding:3px 9px;font-weight:600;">3 · Régler &amp; dessiner ⬇</span>'
    '<span style="background:#e7f0f7;border-radius:6px;padding:3px 9px;">4 · Lancer les calculs</span>'
    '</div>',
    unsafe_allow_html=True,
)
st.caption("Étape 3 — ouvrez un panneau ci-dessous pour régler la distance, filtrer/annoter, ou choisir les profils, puis dessinez sur la carte.")

with st.container():
    # 4 colonnes étroites pour aligner les boutons horizontalement
    c1, c2, c3, c4 = st.columns([1, 1, 1, 1], gap="small")
    st.markdown('<div class="compact-buttons">', unsafe_allow_html=True)

    with c1:
        if st.button(
            "3a · ⚙️ Distance & courbure",
            help="Afficher/masquer le réglage de la méthode de distance, du facteur de courbure et des paramètres d'affichage de la carte",
            use_container_width=True,
            key="btn_dist_curv"
        ):
            ss["dist_panel_open"] = not ss["dist_panel_open"]

    with c2:
        if st.button(
            "3b · 🗂️ Filtrage & légende",
            help="Afficher/masquer les filtres PR (Fait/Ausculte/A_refaire) et le bloc légende",
            use_container_width=True,
            key="btn_filter_legend"
        ):
            ss["legend_panel_open"] = not ss["legend_panel_open"]

    with c3:
        if st.button(
            "3c · ⭕ Cercles & annotations",
            help="Afficher/masquer l'outil d'ajout/suppression des cercles d’annotation",
            use_container_width=True,
            key="btn_circles_ann"
        ):
            ss["circles_panel_open"] = not ss["circles_panel_open"]

    # 👇 Nouveau bouton
    with c4:
        if st.button(
            "3d · 🧩 Profils & éléments",
            help="Afficher/masquer la section 'Profils & éléments à inclure' (et Largeurs)",
            use_container_width=True,
            key="btn_profiles_panel"
        ):
            ss["profile_panel_open"] = not ss["profile_panel_open"]

    st.markdown('</div>', unsafe_allow_html=True)

# Panneau repliable (identique aux options existantes, mais caché par défaut)
if ss["dist_panel_open"]:
    with st.expander("⚙️ Distance & courbure", expanded=True):
        colD, colE, colF, colF2 = st.columns([1.2, 1, 1, 1.2])
        with colD:
            st.selectbox(
                "Méthode de distance",
                ["Segment édité", "Chainage", "Droite PR→PR", "PR × 1000 (fixe)", "Fixe"],
                key="dist_method",
                help=(
                    "Segment édité = distance des polylignes dessinées sur la carte. "
                    "Chainage = delta de chainage_m (ou 1000 m/PR si absent). "
                    "Droite PR→PR = distance droite entre les 2 PR. "
                    "PR × 1000 (fixe) = 1000 m par PR (forcé). "
                    "Fixe = valeur imposée manuellement."
                ),
            )
        with colE:
            st.number_input("Distance fixe (m)", value=float(default_fixed), step=50.0, min_value=0.0, key="fixed_m")
        with colF:
            st.number_input("Facteur de courbure", value=float(ss.get("curvature_factor", 1.00)), step=0.01, min_value=0.90, max_value=1.20, key="curvature_factor")
        with colF2:
            st.slider("Hauteur carte (px)", min_value=500, max_value=1000, value=int(ss.get("map_height", 750)), step=10, key="map_height")
            st.slider("Zoom initial", min_value=10, max_value=22, value=int(ss.get("zoom_init", 14)), step=1, key="zoom_init")

# Variables locales pour la suite du script (comme avant)
dist_method = ss["dist_method"]
fixed_m = float(ss.get("fixed_m", default_fixed))
curvature_factor = float(ss["curvature_factor"])
map_height = int(ss["map_height"])
zoom_init = int(ss["zoom_init"])


# ─────────────────────────────────────────────────────────────
# Profils & éléments (piloté par le bouton "🧩 Profils & éléments")
# ─────────────────────────────────────────────────────────────
st.markdown("---")

if ss["profile_panel_open"]:
    with st.expander("🧩 Profils & éléments à inclure", expanded=True):
        st.markdown("#### Profils et éléments à inclure")

        # 🧭 UX: créer ET modifier des profils (s'ajoutent/écrasent partout les sélecteurs)
        st.markdown("**➕ Créer / ✏️ Modifier un profil**")
        with st.container(border=True):
            # Valeurs par défaut de l'éditeur (sans 'value=' sur les widgets → pas de conflit Streamlit)
            st.session_state.setdefault("new_profile_name", "")
            st.session_state.setdefault("new_profile_elems", ["BDG", "VL", "VR", "BAU"])
            st.session_state.setdefault("new_profile_color", "#e377c2")

            # 1) Charger un profil existant (de base ou perso) dans l'éditeur pour le modifier
            _all_prof_names = list(PROFILES.keys())
            _lc1, _lc2 = st.columns([2, 1])
            with _lc1:
                _load_sel = st.selectbox(
                    "Charger un profil existant pour le modifier",
                    ["— nouveau profil —"] + _all_prof_names,
                    key="load_profile_sel",
                    format_func=lambda k: k.replace("_", " "),
                )
            with _lc2:
                st.write("")
                if st.button("📂 Charger dans l'éditeur", use_container_width=True,
                             key="btn_load_profile", disabled=(_load_sel == "— nouveau profil —")):
                    _src = PROFILES.get(_load_sel, {})
                    st.session_state["new_profile_name"] = _load_sel
                    st.session_state["new_profile_elems"] = [e for e in ALL_ELEMENTS if e in _src]
                    for _e in ALL_ELEMENTS:
                        if _e in _src:
                            st.session_state[f"new_profile_count_{_e}"] = float(_src[_e])
                    st.session_state["new_profile_color"] = PROFILE_COLORS.get(_load_sel, "#e377c2")
                    st.rerun()

            # 2) Éditeur (nom + éléments + comptages + couleur)
            _cpc1, _cpc2 = st.columns([1.4, 1])
            with _cpc1:
                st.text_input("Nom du profil", key="new_profile_name",
                              placeholder="ex. 5_voies, Bretelle_large…")
                _new_prof_elems = st.multiselect(
                    "Éléments composant le profil", ALL_ELEMENTS, key="new_profile_elems",
                    help="Chaque élément vaut 1 voie équivalente par défaut (modifiable ci-dessous).",
                )
            with _cpc2:
                st.color_picker("Couleur du profil", key="new_profile_color")
            if _new_prof_elems:
                _cc = st.columns(min(4, len(_new_prof_elems)))
                for _i, _e in enumerate(_new_prof_elems):
                    st.session_state.setdefault(f"new_profile_count_{_e}", 1.0)
                    with _cc[_i % len(_cc)]:
                        st.number_input(f"{_e} ×", min_value=0.0, step=0.5, key=f"new_profile_count_{_e}")

            # 3) Enregistrer (création si nouveau nom, modification/écrasement si nom existant)
            _bc1, _bc2 = st.columns(2)
            with _bc1:
                if st.button("💾 Enregistrer le profil", use_container_width=True, key="btn_save_profile"):
                    _nm = (st.session_state.get("new_profile_name") or "").strip().replace(" ", "_")
                    _elems = st.session_state.get("new_profile_elems", [])
                    if not _nm:
                        st.warning("Donne un nom au profil.")
                    elif not _elems:
                        st.warning("Sélectionne au moins un élément.")
                    else:
                        st.session_state["custom_profiles"][_nm] = {
                            e: float(st.session_state.get(f"new_profile_count_{e}", 1.0)) for e in _elems
                        }
                        st.session_state["custom_profile_colors"][_nm] = st.session_state.get(
                            "new_profile_color", "#e377c2")
                        _verb = "modifié" if _nm in _BUILTIN_PROFILE_NAMES else "enregistré"
                        st.success(f"Profil « {_nm.replace('_', ' ')} » {_verb} et disponible partout.")
                        st.rerun()
            with _bc2:
                _customs = list(st.session_state.get("custom_profiles", {}).keys())
                _to_del = st.selectbox("Profil perso / modifié à retirer", ["—"] + _customs,
                                       key="del_profile_sel", format_func=lambda k: k.replace("_", " "))
                if st.button("🗑️ Supprimer / réinitialiser", use_container_width=True,
                             key="btn_del_profile", disabled=(_to_del == "—")):
                    st.session_state["custom_profiles"].pop(_to_del, None)
                    st.session_state["custom_profile_colors"].pop(_to_del, None)
                    st.rerun()

            if st.session_state.get("custom_profiles"):
                st.caption(
                    "Profils perso / modifiés actifs : "
                    + ", ".join(n.replace("_", " ") for n in st.session_state["custom_profiles"].keys())
                    + ". Supprimer un profil de base modifié le réinitialise à sa valeur d'origine."
                )
        st.markdown("---")

        colG, colH = st.columns([1.1, 1])
        with colG:
            # Mode édition simple : 1 seul profil, pas de pourcentages
            if dist_method == "Segment édité":
                if ss.get("quick_profile_choice", {}).get(seg_key):
                    st.session_state["profile_simple_select"] = ss["quick_profile_choice"][seg_key]

                profile_simple = st.selectbox(
                    "Profil du sous-segment (mode édition simple — pas de pourcentages)",
                    list(PROFILES.keys()),
                    index=0,
                    key="profile_simple_select",
                    help="En édition, on saisit un seul profil par ligne pour aller plus vite."
                )
                profiles_selected = [profile_simple]
                percents = [100]
                st.caption("Les pourcentages sont masqués en mode 'Segment édité'.")
            else:
                profiles_selected = st.multiselect(
                    "Profils à appliquer (pour le prochain dessin ou le segment global)",
                    list(PROFILES.keys()),
                    default=["2_voies"],
                    help="Tu peux mixer plusieurs profils via des pourcentages (les poids sont normalisés).",
                )
                percents = []
                for name in profiles_selected:
                    perc = st.slider(
                        f"Part du profil {name.replace('_', ' ')} (%)",
                        0, 100, 100 if len(profiles_selected) == 1 else 0, step=5
                    )
                    percents.append(perc)

            profile_mix = merge_profile_mix(profiles_selected, percents)
            if not profile_mix:
                st.warning("Sélectionne au moins un profil avec une part > 0%.")

        dom_name = dominant_profile_name(profiles_selected, percents)
        seg_color = PROFILE_COLORS.get(dom_name, "#ff7f50")

        with colH:
            preset = st.radio("Préréglages d’inclusion", ["Voies", "Tout", "Personnalisé"], index=1, horizontal=True)
            if preset == "Voies":
                included_elements = [e for e in ["VL", "VR", "VM", "VS"] if e in ALL_ELEMENTS]
            elif preset == "Tout":
                included_elements = ALL_ELEMENTS.copy()
            else:
                included_elements = st.multiselect("Éléments inclus", ALL_ELEMENTS, default=["BDG", "VL", "VR", "VM", "BAU", "BRET"])

        # Regrouper "Largeurs" sous ce même panneau
        st.markdown("---")
        # 🧭 UX: conteneur encadré au lieu d'un expander imbriqué (interdit par Streamlit récent)
        st.markdown("**⚙️ Largeurs par élément (m) et surcharges par tronçon**")
        with st.container(border=True):
            widths = {
                e: st.number_input(f"{e}", value=float(DEFAULT_WIDTHS.get(e, 0.0)), step=0.1, min_value=0.0)
                for e in ALL_ELEMENTS
            }
            widths_applied = apply_overrides(widths, overrides, route, cote, float(pr_start), float(pr_end))
            if overrides is not None:
                if widths_applied != widths:
                    st.info("Des surcharges 'overrides' ont été appliquées à ce tronçon.")
                else:
                    st.caption("Aucune surcharge 'overrides' correspondante pour ce tronçon.")
            else:
                widths_applied = widths

        # ⇢ Persistance pour réutilisation quand le panneau est fermé
        ss["profiles_selected"] = profiles_selected
        ss["percents"] = percents
        ss["included_elements"] = included_elements
        ss["widths_applied"] = widths_applied
        ss["dom_name"] = dom_name
        ss["seg_color"] = seg_color

else:
    # Panneau masqué : on réutilise la dernière valeur connue ou des défauts sûrs
    profiles_selected = ss.get("profiles_selected", ["2_voies"])
    percents          = ss.get("percents", [100])
    included_elements = ss.get("included_elements", ALL_ELEMENTS.copy())
    widths_applied    = ss.get("widths_applied", DEFAULT_WIDTHS.copy())
    dom_name          = ss.get("dom_name", dominant_profile_name(profiles_selected, percents))
    seg_color         = ss.get("seg_color", PROFILE_COLORS.get(dom_name, "#ff7f50"))
    profile_mix       = merge_profile_mix(profiles_selected, percents)


# --- PATCH : appliquer le profil choisi via le mini-panneau, même panneau fermé ---
if st.session_state.get("quick_profile_apply", {}).pop(seg_key, False):
    chosen = st.session_state.get("quick_profile_choice", {}).get(seg_key)
    if chosen:
        # garder le selectbox cohérent si on ouvre le panneau plus tard
        st.session_state["profile_simple_select"] = chosen

        # forcer l'état courant du profil utilisé
        profiles_selected = [chosen]
        percents = [100]
        dom_name = chosen
        seg_color = PROFILE_COLORS.get(dom_name, "#ff7f50")
        profile_mix = merge_profile_mix(profiles_selected, percents)

        # persister pour la suite si le panneau est masqué
        ss["profiles_selected"] = profiles_selected
        ss["percents"] = percents
        ss["dom_name"] = dom_name
        ss["seg_color"] = seg_color

        # refermer le mini-panneau
        ss.setdefault("quick_profile_open", {})[seg_key] = False


# NEW: valeurs par défaut quand le panneau est fermé
filter_options = {
    "Fait Oui / Ausculté Oui": ("oui", "oui"),
    "Fait Oui / Ausculté Non": ("oui", "non"),
    "Fait Non / Ausculté Oui": ("non", "oui"),
    "Fait Non / Ausculté Non": ("non", "non"),
}
_selected_filters = list(filter_options.keys())   # par défaut : tout
_show_arefaire_only = False                       # par défaut : non
# ⚡ FLUIDITÉ: préférence d'affichage "clustering" — persistée même panneau fermé (clé widget)
_cluster_points = st.session_state.get("cluster_points_cb", False)

# NEW: si panneau ouvert -> afficher les contrôles et écraser les valeurs par défaut
if ss["legend_panel_open"]:
    with st.expander("🗂️ Filtrage & légende", expanded=True):
        col_f1, col_f2 = st.columns([2, 1])
        with col_f1:
            _selected_filters = st.multiselect(
                "Afficher les statuts PR",
                options=list(filter_options.keys()),
                default=list(filter_options.keys()),
                help="Filtre combiné conforme à la légende PR (bordure = Fait, remplissage = Ausculté)."
            )
        with col_f2:
            _show_arefaire_only = st.checkbox("Uniquement A_refaire = Oui", value=False)
            # ⚡ FLUIDITÉ: regroupe les points proches pour fluidifier les gros fichiers (OFF par défaut)
            _cluster_points = st.checkbox(
                "Regrouper les points (clustering)",
                value=False,
                key="cluster_points_cb",
                help="Regroupe les PR proches en pastilles tant qu'on n'a pas zoomé — très fluide "
                     "sur gros fichiers. Décoché = rendu individuel actuel (popups par point).",
            )

        st.caption("Astuce : décoche une ou plusieurs combinaisons pour synchroniser carte et légende.")

# 2) Préparer le DataFrame DIRMED à partir de df déjà chargé (inchangé)
dirmed_df_all = _prep_dirmed_df(df)
# ⚡ FLUIDITÉ: copie défensive — le résultat est mis en cache et réécrit en place plus bas,
# on travaille donc sur une copie pour ne jamais corrompre l'objet mémorisé par @st.cache_data.
if dirmed_df_all is not None:
    dirmed_df_all = dirmed_df_all.copy()
    # Normalisation 'oui'/'non'
    for c in ["Fait", "Ausculte"]:
        if c in dirmed_df_all.columns:
            dirmed_df_all[c] = dirmed_df_all[c].astype(str).str.strip().str.lower()
        else:
            st.info("ℹ️ La couche DIRMED est inactive : colonnes manquantes (Fait, Ausculte, structure).")
            dirmed_df_all = dirmed_df_all.iloc[0:0]

    # 3) Appliquer le filtre par combinaisons
    if _selected_filters:
        allowed_pairs = set(filter_options[k] for k in _selected_filters)
        mask = dirmed_df_all.apply(
            lambda r: (r.get("Fait", ""), r.get("Ausculte", "")) in allowed_pairs, axis=1
        )
        dirmed_df_all = dirmed_df_all[mask]
    else:
        dirmed_df_all = dirmed_df_all.iloc[0:0]

    # 4) (Option) Restreindre à A_refaire = Oui
    if _show_arefaire_only:
        if "A_refaire" in dirmed_df_all.columns:
            dirmed_df_all = dirmed_df_all[
                dirmed_df_all["A_refaire"].astype(str).str.strip().str.lower().eq("oui")
            ]
        else:
            st.warning("La colonne 'A_refaire' est absente du fichier : le filtre est ignoré.")
else:
    st.info("ℹ️ La couche DIRMED est inactive : colonnes manquantes (Fait, Ausculte, structure) ou coordonnées invalides.")

# 🧭 UX: export TXT des PR actuellement affichés (après filtres Fait/Ausculté/A_refaire)
if dirmed_df_all is not None and not dirmed_df_all.empty:
    _txt_cols = [c for c in ["route", "cote", "pr", "x", "y", "lat", "lon",
                             "Fait", "Ausculte", "A_refaire", "structure"]
                 if c in dirmed_df_all.columns]
    _txt_data = dirmed_df_all[_txt_cols].to_csv(index=False, sep="\t").encode("utf-8")
    _txt_safe = "".join(ch if ch.isalnum() else "_" for ch in f"{route}_{cote}")
    st.download_button(
        f"📄 Exporter les PR affichés (TXT — {len(dirmed_df_all)} points)",
        data=_txt_data,
        file_name=f"PR_affiches_{_txt_safe}.txt",
        mime="text/plain",
        help="Liste tabulée (route, côté, PR, x, y, lat, lon, statuts, structure) des PR visibles sur la carte.",
    )

# -------------------------
# Carte IGN par défaut
# -------------------------
st.markdown("---")
st.markdown("#### Carte IGN")

# Garde-fous sur center/zoom
map_center = midpoint_wgs(coords_wgs)  # doit renvoyer (lat, lon)
if not (isinstance(map_center, (list, tuple)) and len(map_center) == 2):
    st.stop()  # ou lève une Exception explicite
map_center = (float(map_center[0]), float(map_center[1]))

try:
    _zoom = 14 if zoom_init is None else int(zoom_init)
except Exception:
    _zoom = 14

# Créer la map sans tuile par défaut
# ⚡ FLUIDITÉ: prefer_canvas=True → les CircleMarker PR sont dessinés sur un seul <canvas>
# au lieu de milliers de nœuds SVG. Pan/zoom nettement plus fluides sur gros fichiers.
# Popups, tooltips, outils de dessin et fonds de carte restent identiques.
m = folium.Map(location=map_center, zoom_start=_zoom, control_scale=True, tiles=None, prefer_canvas=True)

# IGN - Orthophotos -> défaut
folium.TileLayer(
    tiles=(
        "https://data.geopf.fr/wmts?"
        "SERVICE=WMTS&REQUEST=GetTile&VERSION=1.0.0"
        "&LAYER=ORTHOIMAGERY.ORTHOPHOTOS"
        "&STYLE=normal&FORMAT=image/jpeg&TILEMATRIXSET=PM"
        "&TILEMATRIX={z}&TILEROW={y}&TILECOL={x}"
    ),
    attr="IGN-F/Géoportail",
    name="IGN Orthophotos",
    overlay=False,
    control=True,
    show=True,          # 👈 affichée au chargement et après rerun
    maxNativeZoom=19,
    maxZoom=22
).add_to(m)

# Esri - masquée au chargement
folium.TileLayer(
    tiles="https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}",
    attr="Esri — World Imagery",
    name="Esri Satellite",
    overlay=False,
    control=True,
    show=False,
    maxNativeZoom=19,
    maxZoom=22,
    detectRetina=True
).add_to(m)

# OSM - masquée au chargement
folium.TileLayer(
    tiles="https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png",
    attr="© OpenStreetMap contributors",
    name="OSM Standard",
    overlay=False,
    control=True,
    show=False,
    maxNativeZoom=19,
    maxZoom=22
).add_to(m)

st.markdown("""
<script>
document.addEventListener("DOMContentLoaded", function() {
    // Observer pour renommer les boutons
    const observer = new MutationObserver(() => {
        const finishBtn = document.querySelector('.leaflet-draw-actions a.leaflet-draw-actions-finish');
        const deleteBtn = document.querySelector('.leaflet-draw-actions a.leaflet-draw-actions-remove-last');
        const cancelBtn = document.querySelector('.leaflet-draw-actions a.leaflet-draw-actions-cancel');

        if (finishBtn && deleteBtn && cancelBtn) {
            finishBtn.textContent = 'Valider';
            deleteBtn.textContent = 'Suppr. dernier';
            cancelBtn.textContent = 'Annuler';
            observer.disconnect();
        }
    });

    observer.observe(document.body, { childList: true, subtree: true });
});
</script>

<style>
/* === Agrandir les boutons Leaflet.Draw (barre d'outils) === */
.leaflet-draw-toolbar a {
    width: 88px !important;
    height: 88px !important;
    background-size: 58px 58px !important;
    background-position: center center !important;
    border-radius: 8px !important;
}

/* === Agrandir les boutons d'actions (Finish, Delete, Cancel) === */
.leaflet-draw-actions a {
    font-size: 22px !important;
    padding: 16px 28px !important;
    height: auto !important;
    border-radius: 8px !important;
    background: #fff !important;
    border: 1px solid #ccc !important;
    color: #333 !important;
    text-decoration: none !important;
}

/* Espacement entre les boutons d'action */
.leaflet-draw-actions {
    gap: 12px;
}

/* === Style popup (inchangé) === */
.leaflet-popup-content-wrapper {
  border-radius: 12px !important;
  box-shadow: 0 10px 28px rgba(0,0,0,0.18) !important;
  border: 1px solid #e5e8ed !important;
}
.leaflet-popup-content { margin: 8px 10px !important; }
.leaflet-popup-tip {
  background: #ffffff !important;
  border: 1px solid #e5e8ed !important;
}
</style>
""", unsafe_allow_html=True)


# Marqueurs PR
folium.Marker(coords_wgs[0], tooltip=f"PR {int(pr_start)}", icon=folium.Icon(color="green")).add_to(m)
folium.Marker(coords_wgs[1], tooltip=f"PR {int(pr_end)}", icon=folium.Icon(color="red")).add_to(m)

# Ligne de base PR→PR (bleu clair)
#folium.PolyLine(coords_wgs, color="#6baed6", weight=3, opacity=0.9, tooltip="PR→PR").add_to(m)

# État en session : géométries éditées (temp) et sous-segments (persistants)
if "edited_geoms" not in st.session_state:
    st.session_state["edited_geoms"] = {}  # {seg_key: List[List[(lat,lon)]]}
if "subsegments" not in st.session_state:
    st.session_state["subsegments"] = {}  # {seg_key: [ {"wgs":[...], "mix":{...}, "color":"#hex", "included":[...], "profile_name": str, "profile_label": str} ]}
# ### AJOUT : compteur par profil et par segment pour les labels incrémentaux
if "profile_counts" not in st.session_state:
    st.session_state["profile_counts"] = {}  # {seg_key: {profile_name: count}}
existing_geom = st.session_state["edited_geoms"].get(seg_key, [])
if "edited_polygons" not in st.session_state:
    st.session_state["edited_polygons"] = {}  # {seg_key: [ [ (lat,lon), ... ], ... ] }

if "surfaces" not in st.session_state:
    st.session_state["surfaces"] = {}  # {seg_key: [ { "wgs":[(lat,lon)...], "name":str, "color":"#hex", "area_m2":float, "rabot_h_cm":float, "mats":[{mat,ep}], ... } ]}

# Afficher les sous-segments existants (couleur spécifique)
for it in st.session_state["subsegments"].get(seg_key, []):
    # ### AJOUT : calcul distance et tooltip avec libellé + distance
    l93_tmp = wgs_to_l93(it["wgs"])
    d_tmp = planimetric_distance_l93(l93_tmp)
    d_tmp = max(d_tmp * float(curvature_factor), 0.0)
    label = it.get("profile_label", "Sous-segment")
    folium.PolyLine(
        it["wgs"],
        color=it["color"],
        weight=6,
        opacity=0.70,
        tooltip=f"{label} — {d_tmp:.2f} m"
    ).add_to(m)

# =========================
# SURFACES : polygones déjà enregistrés
# =========================
for surf in st.session_state["surfaces"].get(seg_key, []):
    poly = surf.get("wgs", [])
    if poly:
        folium.Polygon(
            locations=poly,
            color=surf.get("color", "#AA00FF"),
            weight=2,
            fill=True,
            fill_opacity=0.25,
            tooltip=f"{surf.get('name','Surface')} - {surf.get('area_m2',0):.0f} m²"
        ).add_to(m)


# Afficher l'existant temporaire (toutes polylignes) en couleur actuelle du profil
if existing_geom:
    # compat : si ancien format (une seule polyligne), encapsule
    if existing_geom and isinstance(existing_geom[0], tuple):
        existing_geom = [existing_geom]
    for line in existing_geom:
        folium.PolyLine(line, color=seg_color, weight=5, opacity=0.95, tooltip="Segment édité").add_to(m)

# Outil de dessin : polygon toujours, polyline seulement si "Segment édité"
polyline_opts = {"shapeOptions": {"color": seg_color, "weight": 5}} if dist_method == "Segment édité" else False
Draw(
    export=False,
    draw_options={
        "polyline": polyline_opts,  # lignes activées seulement si Segment édité
        "polygon": {"shapeOptions": {"color": "#AA00FF", "fillColor": "#AA00FF", "fillOpacity": 0.20}},
        "rectangle": {"shapeOptions": {"color": "#AA00FF", "fillColor": "#AA00FF", "fillOpacity": 0.15}},
        "circle": False,
        "marker": False,
        "circlemarker": False,
    },
    edit_options={"edit": True, "remove": True},
).add_to(m)

with st.expander("ℹ️ Comment utiliser les outils de dessin (segments & surfaces)", expanded=False):
    st.markdown("""
**A. Saisie rapide d’un sous-segment sur le bouton a gauche "ligne"**
1. **Choisissez le Profil** a droite de la carte (ex. 2_voies, 3_voies, Accès).
3. Cliquez sur **Utiliser ce profil"** pour enregistrer la ligne.
5. Cliquez sur **“➕ Ajouter comme sous-segment”** (colonne de droite) pour l’appliquer au profil choisi.
6. Vous pouvez **modifier les largeurs** de ce sous-segment à tout moment.

**B. Création d’une surface (polygone)**
1. Cliquez l’outil **Polygone** ou **Rectangle** dans la barre de dessin.
2. Dessinez votre surface, puis **Valider**.
3. Dans la colonne de droite, donnez un **nom** puis cliquez **“➕ Ajouter comme surface”**.
4. Retrouvez vos surfaces dans l’onglet **Surface** : saisissez la **hauteur de rabotage (cm)** et vos **matériaux** pour obtenir **volumes** et **tonnages** (export CSV).

**Astuce**
- Les **surfaces affichent l’aire** (m²) arrondie sans décimale.
- Vous pouvez **supprimer** les tracés temporaires (polygones non ajoutés) avec le bouton 🗑️.
""")


# >>> MODIF : légende sans % si mode 'Segment édité'
legend_html = make_legend_html(profiles_selected, percents, show_percentages=(dist_method != "Segment édité"))
m.get_root().html.add_child(folium.Element(legend_html))

# Légende automatique des points PR (bas-gauche)
pr_legend_html = make_pr_points_legend(dirmed_df_all)
m.get_root().html.add_child(folium.Element(pr_legend_html))


# ─────────────────────────────────────────────────────────────
# Couche DIRMED — Ajout des points et des labels de villes
# (Coordonnées WGS84, identiques à celles utilisées sur la carte)
# ─────────────────────────────────────────────────────────────
if dirmed_df_all is not None and not dirmed_df_all.empty:
    layer_dirmed = folium.FeatureGroup(name="Affichage", show=True)

    # ⚡ FLUIDITÉ: cible des points PR principaux.
    # - Option décochée (défaut) → couche directe = comportement actuel inchangé.
    # - Option cochée → MarkerCluster (conserve popups & tooltips de chaque point).
    # Les labels de villes et les anneaux "À refaire" restent toujours sur layer_dirmed.
    if _cluster_points:
        point_target = MarkerCluster(name="PR (regroupés)")
        point_target.add_to(m)
    else:
        point_target = layer_dirmed

    # Villes du 13 (DivIcon texte)
    for city, (clat, clon) in BOUCHES_DU_RHONE_CITIES.items():
        folium.Marker(
            location=(clat, clon),
            icon=DivIcon(
                icon_size=(120, 10),
                icon_anchor=(0, 0),
                # CHANGEMENT : HTML non échappé (vraies balises <div> ... </div>)
                html=f"""
                <div style="
                    font-size:12px;font-weight:700;color:#222;
                    text-shadow: 0 0 3px #ffffff, 0 0 6px #ffffff;
                    background: rgba(255,255,255,0.0); padding: 0 2px;">
                    {city}
                </div>
                """
            ),
            tooltip=city
        ).add_to(layer_dirmed)

    # Points stylés selon Fait / Ausculte (+ anneau si A_refaire=oui)
    for _, r in dirmed_df_all.iterrows():
        fait = _normalize_yn(r.get("Fait", ""))
        ausc = _normalize_yn(r.get("Ausculte", ""))
        aref = _normalize_yn(r.get("A_refaire", ""))  # peut ne pas exister dans tous les fichiers

        sty = PR_STYLE.get((fait, ausc), PR_STYLE[('non', 'non')])

        tooltip = f"{r.get('route','')} - {r.get('pr','')} ({r.get('cote','')})"
        popup_html = build_pr_popup_html(r)  # déjà défini dans ton code
        iframe = IFrame(html=popup_html, width=320, height=210)
        popup = folium.Popup(iframe, max_width=320)

        # marqueur principal
        folium.CircleMarker(
            location=(float(r["lat"]), float(r["lon"])),
            radius=5,  # légèrement plus lisible
            color=sty["stroke"],
            weight=2,
            fill=True,
            fill_color=sty["fill"],
            fill_opacity=1.0 if ausc == 'oui' else 0.6,
            tooltip=tooltip,
            popup=popup
        ).add_to(point_target)  # ⚡ FLUIDITÉ: couche directe ou cluster selon l'option

        # anneau si A_refaire = oui — même popup/tooltip que le marqueur principal.
        # ⚠️ On recrée un IFrame/Popup DÉDIÉS : un objet folium.Popup ne peut pas être
        # attaché à deux marqueurs (cela corromprait le rendu de toute la carte).
        # (avec prefer_canvas=True, l'anneau dessiné en dernier intercepterait le clic
        # et bloquerait le popup du marqueur du dessous sans ce binding)
        if aref == 'oui':
            iframe_ring = IFrame(html=popup_html, width=320, height=210)
            popup_ring = folium.Popup(iframe_ring, max_width=320)
            folium.CircleMarker(
                location=(float(r["lat"]), float(r["lon"])),
                radius=8,
                color="#c0392b",
                weight=2,
                fill=False,
                opacity=0.9,
                tooltip=tooltip,
                popup=popup_ring,
            ).add_to(layer_dirmed)
            


# Un seul LayerControl (replié -> icône en haut-droite)
layer_dirmed.add_to(m) 
folium.LayerControl(collapsed=True, position="topright").add_to(m)

# =========================
# OUTIL : Cercles d'annotation pour PR intermédiaires
# =========================

# (Optionnel) Nettoyage : une seule clé d'état
ss.setdefault("circles_panel_open", False)  # garde "circles_panel_open" comme unique référence
# Affichage conditionnel du panneau en fonction du bouton de la toolbar
if ss["circles_panel_open"]:
    with st.expander("⭕ Cercles & annotations", expanded=True):
        st.markdown("### Ajouter des cercles d'annotation (PR intermédiaires)")

        # Initialisation de la liste des cercles
        if "circles" not in st.session_state:
            st.session_state["circles"] = []

        # Choix du PR de base (début, fin ou autre)
        pr_options = [f"PR début ({pr_start})", f"PR fin ({pr_end})"] + [f"PR {p}" for p in subset["pr"].tolist()]
        selected_pr = st.selectbox("Choisir le PR de base", pr_options)

        # Rayon et nom du PR intermédiaire
        rayon_m = st.number_input("Rayon (m)", min_value=10.0, step=10.0, value=200.0)
        nom_pr = st.text_input("Nom du PR intermédiaire", value=f"{selected_pr} + {int(rayon_m)}")

        # Bouton pour ajouter le cercle
        if st.button("➕ Ajouter ce cercle", key="btn_add_circle"):
            # Trouver les coordonnées du PR choisi
            if "début" in selected_pr:
                base_point = coords_wgs[0]
            elif "fin" in selected_pr:
                base_point = coords_wgs[1]
            else:
                pr_num = float(selected_pr.replace("PR ", ""))
                row = subset[subset["pr"] == pr_num].iloc[0]
                base_point = TO_WGS84.transform(row["x"], row["y"])[::-1]  # (lat, lon)

            st.session_state["circles"].append({
                "center": base_point,
                "radius": rayon_m,
                "label": nom_pr
            })

        # Bouton pour réinitialiser tous les cercles
        if st.button("🗑️ Supprimer tous les cercles", key="btn_clear_circles"):
            st.session_state["circles"] = []

        # Affichage des cercles et annotations sur la carte
        for idx, c in enumerate(st.session_state["circles"]):
            folium.Circle(
                location=c["center"],
                radius=c["radius"],
                color="blue",
                fill=True,
                fill_opacity=0.1,
                tooltip=c["label"]
            ).add_to(m)

            folium.Marker(
                location=c["center"],
                icon=DivIcon(
                    icon_size=(150, 36),
                    icon_anchor=(0, 0),
                    html=f'<div style="font-size:16px;font-weight:bold;color:#003366;">{c["label"]}</div>'
                )
            ).add_to(m)


col_map, col_actions = st.columns([4, 1])
with col_map:
    # Conserver un key stable évite les remounts (optionnel mais recommandé)
    # ⚡ FLUIDITÉ: ne renvoyer que les données de dessin réellement lues par l'app
    # (parse_drawn_polylines / parse_drawn_polygons) → payload allégé, reruns plus rapides.
    map_data = st_folium(
        m, height=map_height, width=None, key=f"map_{seg_key}",
        returned_objects=["last_active_drawing", "all_drawings"],
    ) or {}

# ⬇⬇⬇ PATCH 2+3 — Mise à jour immédiate de l’état des polylignes éditées (avec dédoublonnage)
if dist_method == "Segment édité":
    drawn_list = parse_drawn_polylines(map_data)  # garde ta fonction existante
    if drawn_list:  # on ne touche pas à l'état si aucune nouveauté
        prev = st.session_state["edited_geoms"].get(seg_key, [])
        # Compat : ancien format (une seule ligne) -> encapsuler
        if prev and isinstance(prev[0], tuple):
            prev = [prev]

        # Dédoublonnage tolérant (micro-variations decimales)
        def _as_hash(poly, prec=6):
            return '|'.join(f'{round(lat, prec)},{round(lon, prec)}' for lat, lon in poly)

        seen = { _as_hash(p) for p in prev }
        # On ajoute d'abord la nouveauté de ce run, puis on élimine tout doublon
        merged = prev + [p for p in drawn_list if _as_hash(p) not in seen]

        if merged != prev:  # évite les écritures inutiles (donc évite un rerender gratuit)
            st.session_state["edited_geoms"][seg_key] = merged
            # Polygones (surfaces) dessinés
            polys = parse_drawn_polygons(map_data)
            if polys:
                prevp = st.session_state["edited_polygons"].get(seg_key, [])
                # déduplication
                def _ph(poly, prec=6): return '\n'.join(f"{round(lat,prec)},{round(lon,prec)}" for lat,lon in poly)
                seenp = {_ph(p) for p in prevp}
                mergedp = prevp + [p for p in polys if _ph(p) not in seenp]
                if mergedp != prevp:
                    st.session_state["edited_polygons"][seg_key] = mergedp       
            # >>> AJOUT : ouvrir le panneau rapide pour ce segment
            ss.setdefault("quick_profile_open", {})[seg_key] = True
            ss.setdefault("quick_profile_choice", {}).pop(seg_key, None)
            ss.setdefault("profile_ready", {})[seg_key] = False  # on réinitialise le choix rapide

# Polygones (surfaces) dessinés — hors dépendance aux polylignes
polys_any = parse_drawn_polygons(map_data)
if polys_any:
    prevp = st.session_state["edited_polygons"].get(seg_key, [])
    def _ph2(poly, prec=6):
        return '\n'.join(f"{round(lat,prec)},{round(lon,prec)}" for lat,lon in poly)
    seenp = {_ph2(p) for p in prevp}
    mergedp = prevp + [p for p in polys_any if _ph2(p) not in seenp]
    if mergedp != prevp:
        st.session_state["edited_polygons"][seg_key] = mergedp

with col_actions:
    st.markdown("**Actions**")
    if st.button("Réinitialiser le tracé édité"):
        st.session_state["edited_geoms"].pop(seg_key, None)
        # >>> AJOUT : fermer le mini-panneau si ouvert
        ss.get("quick_profile_open", {}).pop(seg_key, None)
        ss.get("quick_profile_choice", {}).pop(seg_key, None)
        ss.get("profile_ready", {}).pop(seg_key, None)
        st.rerun()

# Purge des polygones temporaires (non ajoutés)
    if st.button("🗑️ Supprimer les surfaces en cours (polygones non ajoutés)"):
        st.session_state["edited_polygons"].pop(seg_key, None)
        st.rerun()

    # Ajouter comme surface depuis les polygones dessinés
    edited_polys = st.session_state.get("edited_polygons", {}).get(seg_key, []) or []
    if edited_polys:
        # choisir le plus grand polygone par aire projetée
        areas = [polygon_area_m2_from_wgs(p) for p in edited_polys]
        idx_max = int(max(range(len(areas)), key=lambda i: areas[i])) if areas else 0

        # nom par défaut
        nb = len(st.session_state.get("surfaces", {}).get(seg_key, [])) + 1
        surf_name = st.text_input("Nom de la surface", value=f"Surface_{nb}", key=f"surf_name_{seg_key}")

        if st.button("➕ Ajouter comme surface", key=f"btn_add_surface_{seg_key}"):
            chosen_poly = edited_polys[idx_max]
            area_m2 = polygon_area_m2_from_wgs(chosen_poly)
            new_surf = {
                "wgs": chosen_poly,
                "name": surf_name or f"Surface_{nb}",
                "color": "#AA00FF",
                "area_m2": float(area_m2),
                "rabot_h_cm": 0.0,
                "mats": []
            }
            lst = st.session_state["surfaces"].get(seg_key, [])
            lst.append(new_surf)
            st.session_state["surfaces"][seg_key] = lst

            # vider les polygones temporaires pour éviter les doublons
            st.session_state["edited_polygons"].pop(seg_key, None)
            st.rerun()
    else:
        st.caption("Dessine un polygone puis ajoute-le comme surface.")

    # Liste et gestion des sous-segments
    subsegs = st.session_state["subsegments"].get(seg_key, [])
    if subsegs:
        if st.button("🗑️ Supprimer tous les sous-segments"):
            st.session_state["subsegments"][seg_key] = []
            # On ne réinitialise pas les compteurs pour conserver l'historique de numérotation
            st.rerun()

    if dist_method == "Segment édité":
        st.caption("🔶 Mode édition actif : dessine/édite une ou plusieurs polylignes.")
        # Activation ajout sous-segment si au moins une polyligne est présente
        edited_wgs_list = st.session_state["edited_geoms"].get(seg_key) or []
        if edited_wgs_list and isinstance(edited_wgs_list[0], tuple):
            edited_wgs_list = [edited_wgs_list]  # compat ancien format

        # 🧭 UX (F3): flux simplifié — dessiner → choisir le profil → UN seul bouton "Ajouter".
        # Plus de bouton intermédiaire "Utiliser ce profil" ni de bouton désactivé en attente.
        # Le radio est présélectionné sur un vrai profil (le profil courant ou le 1er disponible).
        _prof_keys = list(PROFILES.keys())
        if edited_wgs_list:
            _default_prof = profiles_selected[0] if profiles_selected else _prof_keys[0]
            _default_idx = _prof_keys.index(_default_prof) if _default_prof in _prof_keys else 0
            quick_choice = st.radio(
                "Profil de ce tracé",
                _prof_keys,
                index=_default_idx,
                format_func=lambda k: k.replace("_", " "),
                key=f"quick_prof_radio_{seg_key}",
                horizontal=False,
            )
        else:
            quick_choice = None
            st.caption("✏️ Dessine une polyligne sur la carte pour activer l'ajout d'un sous-segment.")

        # Un seul bouton "Ajouter" : actif dès qu'une polyligne est dessinée.
        if st.button(
            "➕ Ajouter comme sous-segment",
            disabled=not bool(edited_wgs_list),
            key=f"btn_add_subseg_{seg_key}"
        ):
            # 🧭 UX (F3): le profil vient directement du radio ci-dessus (au moment du clic).
            _add_profiles = [quick_choice] if quick_choice else (profiles_selected or ["2_voies"])
            _add_percents = [100] if quick_choice else (percents or [100])

            # --- Choix de la ligne à ajouter ---
            if len(edited_wgs_list) == 1:
                chosen = edited_wgs_list[0]
            else:
                # Prend la plus longue, plus robuste que [-1]
                chosen = max(
                    edited_wgs_list,
                    key=lambda p: planimetric_distance_l93(wgs_to_l93(p))
                )

            # Profil dominant & label incrémental par profil
            dom_for_label = dominant_profile_name(_add_profiles, _add_percents) or "mix"
            counts_map = st.session_state["profile_counts"].get(seg_key, {})
            next_idx = int(counts_map.get(dom_for_label, 0)) + 1
            profile_label = f"{dom_for_label}_{next_idx}"
            counts_map[dom_for_label] = next_idx
            st.session_state["profile_counts"][seg_key] = counts_map

            new_item = {
                "wgs": chosen,
                "mix": merge_profile_mix(_add_profiles, _add_percents),
                "color": PROFILE_COLORS.get(dominant_profile_name(_add_profiles, _add_percents), "#ff7f50"),
                "included": included_elements.copy(),
                # Infos de profil pour libellé stable
                "profile_name": dom_for_label,
                "profile_label": profile_label,
                # Snapshot des largeurs utilisées au moment de l'ajout
                "widths": widths_applied.copy(),
            }

            # Enregistrer le sous-segment dans l'état
            sub_list = st.session_state["subsegments"].get(seg_key, [])
            sub_list.append(new_item)
            st.session_state["subsegments"][seg_key] = sub_list
            # Vider le tracé temporaire pour éviter les doublons au prochain ajout
            st.session_state["edited_geoms"].pop(seg_key, None)
            st.rerun()
    else:
        st.caption("ℹ️ Mode lecture seule (outils masqués).")


    # Liste des sous-segments saisis (résumé compact + détails)
    sub_list = st.session_state["subsegments"].get(seg_key, [])
    if sub_list:
        st.markdown("**Sous-segments saisis**")

        # --- Styles compacts en chips (pas d'élargissement de colonne)
        st.markdown("""
        <style>
        .ss-wrap{display:flex;flex-wrap:wrap;gap:6px;row-gap:6px;align-items:center;max-width:100%;}
        .ss-chip{display:inline-flex;align-items:center;gap:6px;padding:4px 8px;
                border-radius:999px;border:1px solid #e5e7eb;background:#f8fafc;
                font-size:12px;line-height:1;white-space:nowrap;max-width:100%;
                overflow:hidden;text-overflow:ellipsis;}
        .ss-dot{width:8px;height:8px;border-radius:50%;}
        </style>
        """, unsafe_allow_html=True)

        # --- Résumé ultra-compact : Label (couleur) + distance
        st.markdown('<div class="ss-wrap">', unsafe_allow_html=True)
        for idx, it in enumerate(sub_list):
            l93 = wgs_to_l93(it["wgs"])
            dist_m = max(planimetric_distance_l93(l93) * float(curvature_factor), 0.0)
            label = it.get("profile_label", f"{it.get('profile_name','mix')}_{idx+1}")
            st.markdown(
                f'<span class="ss-chip"><span class="ss-dot" style="background:{it["color"]}"></span>'
                f'{label}&nbsp;•&nbsp;{dist_m:.0f}&nbsp;m</span>',
                unsafe_allow_html=True
            )
        st.markdown('</div>', unsafe_allow_html=True)

        # --- Détails/édition (compactés dans un expander)
        with st.expander("Détails & édition (optionnel)", expanded=False):
            to_delete = None
            for idx, it in enumerate(sub_list):
                l93 = wgs_to_l93(it["wgs"])
                dist = max(planimetric_distance_l93(l93) * float(curvature_factor), 0.0)
                label = it.get("profile_label", f"{it.get('profile_name','mix')}_{idx+1}")

                # Compat : initialiser "widths" si ancien sous-segment
                if "widths" not in it or not isinstance(it["widths"], dict):
                    it["widths"] = widths_applied.copy()

                # Carte info compacte (width safe)
                st.markdown(
                    f"""
                    <div style="border:2px solid {it['color']};
                                background-color:{it['color']}22;
                                border-radius:8px;padding:8px;margin-bottom:6px;max-width:100%;">
                    <b>#{idx+1}</b> — <span style="color:{it['color']};font-weight:600;">{label}</span>
                    &nbsp;•&nbsp;{dist:.1f} m
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                # Éditeur des largeurs — conteneur encadré (expander imbriqué interdit)
                st.markdown(f"**✏️ Modifier les largeurs — {label}**")
                with st.container(border=True):
                    st.caption("Ces largeurs n'affectent que ce sous-segment.")
                    cols = st.columns(3)
                    for j, e in enumerate(ALL_ELEMENTS):
                        with cols[j % 3]:
                            current_val = float(it["widths"].get(e, DEFAULT_WIDTHS.get(e, 0.0)))
                            it["widths"][e] = st.number_input(
                                f"{e}", value=current_val, step=0.1, min_value=0.0,
                                key=f"w_{seg_key}_{idx}_{e}"
                            )

                    # 🛠️ BUGFIX: colonnes aplaties (Streamlit interdit l'imbrication de colonnes
                    # sur 2 niveaux : col_actions > m_c > b1/b2 plantait). Deux boutons sur une
                    # seule rangée de colonnes (1 seul niveau d'imbrication dans col_actions).
                    b1, b2 = st.columns(2)
                    with b1:
                        if st.button("↺ Copier global", key=f"copy_global_{seg_key}_{idx}",
                                     use_container_width=True):
                            it["widths"] = widths_applied.copy()
                            st.rerun()
                    with b2:
                        if st.button("↺ Réinit défauts", key=f"reset_defaults_{seg_key}_{idx}",
                                     use_container_width=True):
                            it["widths"] = {e: float(DEFAULT_WIDTHS.get(e, 0.0)) for e in ALL_ELEMENTS}
                            st.rerun()

                # Suppression unitaire – petit bouton
                if st.button(f"Supprimer #{idx+1}", key=f"del_{seg_key}_{idx}"):
                    to_delete = idx

            if to_delete is not None:
                st.session_state["subsegments"][seg_key].pop(to_delete)
                st.rerun()



# =========================
# Calcul des distances et surfaces
# =========================

st.markdown("#### Lancement des calculs")

auto_compute = st.toggle(
    "Calcul automatique à chaque changement", value=False,
    help="Quand désactivé, les calculs ne s'exécutent que via le bouton ci-dessous.",
    key="toggle_auto_compute"
)

# Bouton de calcul propre au segment
button_key = f"compute_btn_{seg_key}"
manual_compute = st.button("🚀 Lancer les calculs", type="primary", key=button_key)

# Petit cache persistant des résultats et état 'débloqué' du panneau Résultats
calc_cache = st.session_state.setdefault("calc_cache", {})  # {seg_key: {"areas_df_fr": df}}
if "calc_unlocked" not in st.session_state:
    st.session_state["calc_unlocked"] = False

# Décision de calcul
do_compute = auto_compute or manual_compute
if do_compute:
    # Dès qu'on calcule une fois, on 'débloque' l'affichage des Résultats pour ce run et les suivants
    st.session_state["calc_unlocked"] = True


subsegs = st.session_state["subsegments"].get(seg_key, [])
areas_df = pd.DataFrame([])
total_area = 0.0
distance_display_m = 0.0  # pour l'affichage métrique en haut

# Chemin 1 : sous-segments présents -> calcul par ligne (profils distincts)
if dist_method == "Segment édité" and subsegs:
    rows_all = []
    total_area_all = 0.0
    total_dist_all = 0.0
    for ss_idx, it in enumerate(subsegs, start=1):
        l93 = wgs_to_l93(it["wgs"])
        d = planimetric_distance_l93(l93)
        d = max(d * float(curvature_factor), 0.0)
        # Utiliser les largeurs propres à ce sous-segment si disponibles
        widths_this = it.get("widths", widths_applied)
        df_part, area_part = compute_areas(d, widths_this, it["mix"], it["included"])
        # ### AJOUT : colonnes distance & libellé profil incrémental
        df_part["__ss__"] = ss_idx
        df_part["distance_m"] = d
        df_part["profil_nom"] = it.get("profile_label", f"{it.get('profile_name','mix')}_{ss_idx}")
        rows_all.append(df_part)
        total_area_all += area_part
        total_dist_all += d
    areas_df = pd.concat(rows_all, ignore_index=True) if rows_all else pd.DataFrame([])
    total_area = float(total_area_all)
    distance_display_m = float(total_dist_all)

# Chemin 2 : comportement global (profil global + 1 géométrie ou chainage/droite/fixe)
else:
    edited_wgs = st.session_state["edited_geoms"].get(seg_key)
    if edited_wgs and isinstance(edited_wgs[0], tuple):
        edited_wgs = [edited_wgs]
    edited_l93 = [wgs_to_l93(line) for line in edited_wgs] if edited_wgs else None
    straight_l93 = coords_l93

    if dist_method == "Segment édité" and edited_l93:
        distance_m = sum(planimetric_distance_l93(line) for line in edited_l93)
    elif dist_method == "Chainage":
        if pd.notna(pr1.get("chainage_m", np.nan)) and pd.notna(pr2.get("chainage_m", np.nan)):
            distance_m = float(pr2["chainage_m"] - pr1["chainage_m"])
        elif pd.notna(pr1["pr"]) and pd.notna(pr2["pr"]):
            distance_m = 1000.0 * float(pr2["pr"] - pr1["pr"])
        else:
            distance_m = planimetric_distance_l93(straight_l93)
    elif dist_method == "Droite PR→PR":
        distance_m = planimetric_distance_l93(straight_l93)
    # >>> MODIF : nouveau cas 'PR × 1000 (fixe)'
    elif dist_method == "PR × 1000 (fixe)":
        distance_m = pr_delta_m(pr_start, pr_end)
    else:  # Fixe
        distance_m = float(fixed_m)

    distance_m = max(distance_m * float(curvature_factor), 0.0)
    areas_df, total_area = compute_areas(distance_m, widths_applied, profile_mix, included_elements)
    distance_display_m = float(distance_m)
    # ### AJOUT : colonnes distance & libellé profil (global)
    global_label = (dominant_profile_name(profiles_selected, percents) or "mix") + "_1"
    areas_df["distance_m"] = distance_m
    areas_df["profil_nom"] = global_label

# ---- Traduction FR des colonnes pour l'affichage
areas_df_fr = areas_df.rename(columns={
    "element": "élément",
    "count_equiv": "comptage_équivalent",
    "width_m": "largeur_m",
    "width_equiv_m": "largeur_équivalente_m",
    "area_m2": "surface_m2",
    "__ss__": "sous_segment",
    # AJOUT :
    "distance_m": "distance_m",
    "profil_nom": "profil_nom",
})

# Réordonner pour lisibilité
cols_order = [
    c for c in ["sous_segment", "profil_nom", "distance_m", "élément",
                "comptage_équivalent", "largeur_m", "largeur_équivalente_m", "surface_m2"]
    if c in areas_df_fr.columns
]
areas_df_fr = areas_df_fr[cols_order + [c for c in areas_df_fr.columns if c not in cols_order]]

# -- Cache du dernier résultat pour ce segment
st.session_state["calc_cache"][seg_key] = {
    "areas_df_fr": areas_df_fr
}
# -- Si on n'a PAS calculé cette fois-ci, réutiliser le dernier résultat si disponible
if not do_compute:
    if st.session_state.get("calc_unlocked") and seg_key in st.session_state["calc_cache"]:
        areas_df_fr = st.session_state["calc_cache"][seg_key]["areas_df_fr"]
        st.info("Affichage des résultats du dernier calcul (pas de mise à jour automatique).")
        # distance_display_m, total_area, recap_elements dépendent de areas_df_fr -> on les reconstitue si besoin
        try:
            # Recalcule minimal pour les KPI d'en-tête à partir du tableau détaillé
            distance_display_m = float(areas_df_fr.get("distance_m", pd.Series([0])).dropna().iloc[0]) if "distance_m" in areas_df_fr.columns else 0.0
            total_area = float(areas_df_fr["surface_m2"].sum()) if "surface_m2" in areas_df_fr.columns else 0.0
            recap_elements = (
                areas_df_fr.groupby("élément", as_index=False)[["surface_m2"]]
                .sum()
                .sort_values("surface_m2", ascending=False)
                if "élément" in areas_df_fr.columns and "surface_m2" in areas_df_fr.columns
                else pd.DataFrame(columns=["élément","surface_m2"])
            )
            surface_totale_voirie = total_area
        except Exception:
            pass  # on garde les valeurs existantes si tout est déjà en place
    else:
        st.info("Ajuste les filtres et la sélection (route/côté/PR), puis lance les calculs.")
        st.stop()
# =========================
# Résultats
# =========================
st.markdown("---")
st.markdown("#### Résultats")
topA, topB, topC, topD = st.columns(4)
with topA:
    st.metric("Distance (m)", f"{distance_display_m:,.0f}".replace(",", " "))
with topB:
    st.metric("Surface totale (m²)", f"{total_area:,.0f}".replace(",", " "))
with topC:
    st.write("Méthode :", dist_method)
with topD:
    st.write("Profil dominant :", (dominant_profile_name(profiles_selected, percents) or "mix / non défini").replace("_", " "))

# Tableau détaillé (avec distances & libellés)
st.dataframe(areas_df_fr, use_container_width=True)

# #### ✅ Récapitulatif global (voirie & éléments)
st.markdown("#### ✅ Récapitulatif global (voirie & éléments)")
surface_totale_voirie = float(areas_df_fr["surface_m2"].sum()) if not areas_df_fr.empty else 0.0
st.write(f"**Surface totale voirie : {surface_totale_voirie:,.0f} m²**".replace(",", " "))

recap_elements = (
    areas_df_fr.groupby("élément", as_index=False)[["surface_m2"]]
    .sum()
    .sort_values("surface_m2", ascending=False)
    if "élément" in areas_df_fr.columns and not areas_df_fr.empty else pd.DataFrame(columns=["élément", "surface_m2"])
)
st.dataframe(_round_df0(recap_elements, _DEC_EXCLUDE), use_container_width=True)

# 🧭 UX: démarre le classeur Excel global de ce run (les onglets ci-dessous le complètent)
st.session_state["export_bundle"] = {"Détail": areas_df_fr, "Récap éléments": recap_elements}

# Récapitulatif par sous-segment (optionnel)
if "sous_segment" in areas_df_fr.columns and not areas_df_fr.empty:
    with st.expander("Récapitulatif par sous-segment", expanded=False):
        recap = (
            areas_df_fr.groupby("sous_segment", as_index=False)[["surface_m2"]]
            .sum()
            .sort_values("sous_segment")
        )
        st.dataframe(_round_df0(recap, _DEC_EXCLUDE), use_container_width=True)
        # Export CSV
        csv_bytes = areas_df_fr.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Télécharger le détail (CSV)",
            data=csv_bytes,
            file_name="sous_segments_detail.csv",
            mime="text/csv",
        )
        csv_bytes2 = recap.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Télécharger le récap par sous-segment (CSV)",
            data=csv_bytes2,
            file_name="sous_segments_recap.csv",
            mime="text/csv",
        )
else:
    # Même sans sous-segments, proposer l’export du détail et du récap éléments
    csv_bytes = areas_df_fr.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Télécharger le détail (CSV)",
        data=csv_bytes,
        file_name="detail.csv",
        mime="text/csv",
    )
    csv_bytes_el = recap_elements.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Télécharger le récap éléments (CSV)",
        data=csv_bytes_el,
        file_name="elements_recap.csv",
        mime="text/csv",
    )






# =========================
# Rabotage & Reprofilage  
# =========================
st.markdown("---")
st.markdown("## Rabotage & Reprofilage & Surface")

# ---- Utilitaires déjà présents plus haut ----
# _ensure_surfaces_source(mode_base, recap_elements_df, surface_totale)
# _select_elements_block(source_df, key_prefix)
# _export_suffix(route, cote, pr_start, pr_end)
# _safe_default_thickness(mat, materials_df)
# _safe_default_density(mat, materials_df)

# --- Sécuriser l'état persistant utilisé par les onglets ---
st.session_state.setdefault("rabot_epaisseurs", {})          # {"VL": 3.0, ...}
st.session_state.setdefault("materials_df", pd.DataFrame(DEFAULT_MATERIALS))
st.session_state.setdefault("reprof_thk_matrix", pd.DataFrame())  # matrice élément×matériau




# --- HOTFIX : utilitaires Rabotage & Reprofilage (réinsérés) ---

def _ensure_surfaces_source(mode_base: str,
                            recap_elements_df: pd.DataFrame,
                            surface_totale: float) -> pd.DataFrame:
    """
    Retourne un DF des surfaces selon la base :
    - 'Toute la voirie' : 1 ligne synthétique (TOUTE_VOIRIE)
    - 'Par élément' : recap_elements (élément + surface_m2)
    """
    if mode_base == "Toute la voirie":
        return pd.DataFrame([{"élément": "TOUTE_VOIRIE", "surface_m2": surface_totale}])
    # Par élément
    df = recap_elements_df.copy()
    if df.empty:
        return pd.DataFrame([{"élément": "(aucun)", "surface_m2": 0.0}])
    return df


def _select_elements_block(source_df: pd.DataFrame, key_prefix: str) -> list[str]:
    """
    UI commune : multi‑sélection des éléments à inclure dans les calculs.
    - source_df : DataFrame contenant au moins ['élément','surface_m2']
    - key_prefix : 'rabot' ou 'reprof' pour isoler l'état Streamlit
    """
    opts = source_df["élément"].astype(str).tolist()
    default = opts  # tout coché par défaut
    st.markdown("**Éléments à inclure**")
    sel = st.multiselect("Éléments", opts, default=default, key=f"{key_prefix}_elems")
    return sel


def _export_suffix(route: str, cote: str, pr_start: float, pr_end: float) -> str:
    def _slug(s: str) -> str:
        s = str(s).strip().replace(" ", "_")
        allowed = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-"
        return "".join(ch if ch in allowed else "_" for ch in s)
    try:
        _r = _slug(route)
        _c = _slug(cote)
        _a = int(float(pr_start))
        _b = int(float(pr_end))
        return f"{_r}_{_c}_PR{_a}-{_b}"
    except Exception:
        return "export"



tab_rabot, tab_reprof, tab_surface, tab_cout = st.tabs(["Rabotage", "Reprofilage", "Surface", "Coût (€)"])

# ─────────────────────────────────────────────────────────────
# Onglet 1 : RABOTAGE  → Surface (m²) × Épaisseur (cm) = Volume (m³)
# ─────────────────────────────────────────────────────────────



with tab_rabot:
    st.subheader("Rabotage (multi-hauteurs)")

    # 1) Base de calcul → source des surfaces
    base_rabot = st.radio(
        "Base de calcul",
        ["Toute la voirie", "Par élément"],
        horizontal=True,
        key=f"base_rabot_multi__{seg_key}",
        help=("Toute la voirie : un total unique. Par élément : un total par BAU/BDG/VL/VR/VM/VS/BRET.")
    )
    rabot_src = _ensure_surfaces_source(base_rabot, recap_elements, surface_totale_voirie).copy()

    # 2) Sélection d’éléments (harmonisée)
    elems_sel_rabot = _select_elements_block(rabot_src, "rabot_multi")
    if elems_sel_rabot:
        rabot_src = rabot_src[rabot_src["élément"].isin(elems_sel_rabot)].copy()


    # 3) Édition multi-hauteurs (passes) par élément
    # Chaque élément a une liste dynamique en session : rabot_list_{seg_key}__{el}
    rows = []
    vol_total_rabot = 0.0

    # >>> NOUVEAU : mémo local pour "Reprendre depuis l'élément au-dessus"
    prev_el_name = None
    prev_el_passes = None

    for _, row in rabot_src.iterrows():
        el = str(row["élément"])
        surf = float(row["surface_m2"]) if pd.notna(row["surface_m2"]) else 0.0
        st.markdown(f"### {el} — {surf:,.0f} m²".replace(",", " "))

        key_prefix = f"{seg_key}__{el}"
        list_key = f"rabot_list_{key_prefix}"

        # Initialisation de la liste des passes (migration depuis rabot_epaisseurs si existant)
        if list_key not in st.session_state or not isinstance(st.session_state[list_key], list) or len(st.session_state[list_key]) == 0:
            migrated = None
            try:
                if base_rabot == "Par élément" and el != "TOUTE_VOIRIE":
                    prev = st.session_state.get("rabot_epaisseurs", {}).get(el)
                    if prev is not None:
                        migrated = [{"label": "Passe héritée", "h": float(prev)}]
            except Exception:
                migrated = None
            st.session_state[list_key] = migrated or [{"label": "Passe 1", "h": 0.0}]

        passes = st.session_state[list_key]

        # --- NOUVEAUTÉ (comportement strict) :
        #     Bouton pour "Reprendre les passes depuis l’élément au-dessus"
        if base_rabot == "Par élément" and prev_el_name is not None:
            with st.expander("Reprendre les passes depuis l’élément au‑dessus", expanded=False):
                if st.button(
                    f"⬇️ Copier depuis {prev_el_name}",
                    key=f"rabot_copy_prev_{key_prefix}",
                    use_container_width=True,
                ):
                    st.session_state[list_key] = [
                        {"label": str(p.get("label", f"Passe {i+1}")), "h": float(p.get("h", 0.0))}
                        for i, p in enumerate(prev_el_passes or [])
                    ]
                    st.rerun()
        # --- fin nouveauté

        # Actions rapides pour l'élément : mise à jour en masse des hauteurs
        with st.expander("Mise à jour rapide des hauteurs pour cet élément", expanded=False):
            c_mass1, c_mass2 = st.columns([2, 1])
            with c_mass1:
                new_h = st.number_input(
                    f"Hauteur commune (cm) pour {el}",
                    min_value=0.0,
                    step=0.5,
                    value=0.0,
                    key=f"rabot_mass_val_{key_prefix}",
                )
            with c_mass2:
                if st.button(
                    f"Appliquer à toutes les passes de {el}",
                    key=f"rabot_mass_apply_{key_prefix}",
                    use_container_width=True,
                ):
                    for i in range(len(passes)):
                        passes[i]["h"] = float(new_h)
                    st.rerun()

        # Lignes dynamiques : label + hauteur (cm) + suppression
        for idx, item in enumerate(list(passes)):
            col1, col2, col3 = st.columns([2, 2, 1], vertical_alignment="center")
            with col1:
                label = st.text_input(
                    f"Nom passe {idx+1}",
                    value=str(item.get("label", f"Passe {idx+1}")),
                    key=f"rab_lbl_{key_prefix}_{idx}",
                )
            with col2:
                h = st.number_input(
                    "Hauteur (cm)",
                    min_value=0.0,
                    step=0.5,
                    value=float(item.get("h", 0.0)),
                    key=f"rab_h_{key_prefix}_{idx}",
                )
            with col3:
                if st.button("❌", key=f"rab_del_{key_prefix}_{idx}"):
                    passes.pop(idx)
                    st.rerun()

            # MàJ état
            passes[idx]["label"] = label
            passes[idx]["h"] = h

            # Calcul m³ pour cette passe
            vol = surf * (h / 100.0)  # m³ = m² × (cm/100)
            rows.append(
                {
                    "élément": el,
                    "passe": label,
                    "surface_m2": surf,
                    "hauteur_cm": h,
                    "volume_m3": vol,
                }
            )
            vol_total_rabot += vol

        # Boutons d’action par élément : ajouter / réinitialiser
        c_add, c_reset = st.columns([1, 1])
        with c_add:
            if st.button(f"+ Ajouter une passe pour {el}", key=f"rab_add_{key_prefix}", type="secondary"):
                passes.append({"label": f"Passe {len(passes)+1}", "h": 0.0})
                st.rerun()
        with c_reset:
            if st.button(f"⟲ Réinitialiser {el}", key=f"rab_reset_{key_prefix}"):
                st.session_state[list_key] = [{"label": "Passe 1", "h": 0.0}]
                st.rerun()

        # >>> Mémoriser l’élément courant comme "précédent" pour le suivant
        prev_el_name = el
        prev_el_passes = [
            {"label": str(p.get("label", f"Passe {i+1}")), "h": float(p.get("h", 0.0))}
            for i, p in enumerate(st.session_state[list_key] or [])
        ]

    # 4) Résultats & Exports
    df_rabot = pd.DataFrame(rows)
    st.markdown("### ✅ Détail rabotage (multi-hauteurs)")
    if not df_rabot.empty:
        view_cols = ["élément", "passe", "surface_m2", "hauteur_cm", "volume_m3"]
        st.dataframe(_round_df0(df_rabot[view_cols], _DEC_EXCLUDE), use_container_width=True)
        # 🧭 UX: ajoute le détail rabotage au classeur Excel global
        st.session_state.setdefault("export_bundle", {})["Rabotage"] = df_rabot[view_cols]

        # KPIs
        k1, k2 = st.columns(2)
        with k1:
            st.metric("Surface totale (sélection) m²", f"{df_rabot['surface_m2'].sum():,.0f}".replace(",", " "))
        with k2:
            st.metric("Volume total rabotage (m³)", f"{vol_total_rabot:,.0f}".replace(",", " "))

        # Totaux par élément
        st.markdown("#### Totaux par élément")
        tot_el = (
            df_rabot.groupby("élément", as_index=False)[["volume_m3"]]
            .sum()
            .sort_values("volume_m3", ascending=False)
        )
        st.dataframe(_round_df0(tot_el, _DEC_EXCLUDE), use_container_width=True)

        # 👉 Suffixe d'export (DÉFINI AVANT tout usage)
        _suf = _export_suffix(route, cote, pr_start, pr_end)

        # ── Totaux par hauteur de rabotage (cm)
        st.markdown("#### Totaux par hauteur de rabotage (cm)")
        # Arrondir légèrement pour éviter des doublons 3.0000001, etc.
        tmp_rabot = df_rabot.copy()
        tmp_rabot["hauteur_cm"] = tmp_rabot["hauteur_cm"].round(2)
        recap_hauteurs = (
            tmp_rabot.groupby("hauteur_cm", as_index=False)[["surface_m2", "volume_m3"]]
            .sum()
            .sort_values("hauteur_cm", ascending=True)
        )
        st.dataframe(_round_df0(recap_hauteurs, _DEC_EXCLUDE), use_container_width=True)

        # Export CSV : totaux par hauteur
        st.download_button(
            "Télécharger totaux par hauteur (CSV)",
            data=recap_hauteurs.to_csv(index=False).encode("utf-8"),
            file_name=f"rabotage_totaux_par_hauteur_{_suf}.csv",
            mime="text/csv",
        )

        # ── Cumul progressif par hauteur (tri croissant)
        st.markdown("#### Cumul progressif par hauteur (ordre croissant)")
        recap_hauteurs_cum = recap_hauteurs.copy()
        recap_hauteurs_cum["surface_cumulée_m2"] = recap_hauteurs_cum["surface_m2"].cumsum()
        recap_hauteurs_cum["volume_cumulé_m3"]  = recap_hauteurs_cum["volume_m3"].cumsum()
        st.dataframe(
            _round_df0(recap_hauteurs_cum[
                ["hauteur_cm", "surface_m2", "volume_m3", "surface_cumulée_m2", "volume_cumulé_m3"]
            ], _DEC_EXCLUDE),
            use_container_width=True
        )

        # Export CSV : cumul progressif
        st.download_button(
            "Télécharger cumul par hauteur (CSV)",
            data=recap_hauteurs_cum.to_csv(index=False).encode("utf-8"),
            file_name=f"rabotage_cumul_par_hauteur_{_suf}.csv",
            mime="text/csv",
        )

        # Exports CSV existants
        st.download_button(
            "Télécharger le détail (CSV)",
            data=df_rabot[view_cols].to_csv(index=False).encode("utf-8"),
            file_name=f"rabotage_multi_detail_{_suf}.csv",
            mime="text/csv",
        )
        st.download_button(
            "Télécharger totaux par élément (CSV)",
            data=tot_el.to_csv(index=False).encode("utf-8"),
            file_name=f"rabotage_multi_totaux_par_element_{_suf}.csv",
            mime="text/csv",
        )
    else:
        st.info("Aucune passe de rabotage saisie pour cette sélection.")






# --- Helpers densité / épaisseur par défaut pour les matériaux ---

def _safe_default_thickness(mat: str, materials_df: pd.DataFrame) -> float:
    """
    Retourne l'épaisseur par défaut (cm) du matériau 'mat'.
    Si non trouvée/NA -> 0.0
    """
    if materials_df is None or materials_df.empty:
        return 0.0
    ser = materials_df.loc[
        materials_df["matériau"].astype(str) == str(mat),
        "épaisseur_cm"
    ]
    if ser.empty:
        return 0.0
    try:
        val = ser.iloc[0]
        return float(val) if pd.notna(val) else 0.0
    except Exception:
        return 0.0


def _safe_default_density(mat: str, materials_df: pd.DataFrame) -> float:
    """
    Retourne la densité (t/m³) du matériau 'mat'.
    Si non trouvée/NA -> 0.0
    """
    if materials_df is None or materials_df.empty:
        return 0.0
    ser = materials_df.loc[
        materials_df["matériau"].astype(str) == str(mat),
        "densité_t_m3"
    ]
    if ser.empty:
        return 0.0
    try:
        val = ser.iloc[0]
        return float(val) if pd.notna(val) else 0.0
    except Exception:
        return 0.0


# 🧭 UX: prix unitaire (€/t) d'un matériau depuis la bibliothèque (0.0 si absent/NA)
def _safe_default_price(mat: str, materials_df: pd.DataFrame) -> float:
    if materials_df is None or materials_df.empty or "prix_eur_t" not in materials_df.columns:
        return 0.0
    ser = materials_df.loc[materials_df["matériau"].astype(str) == str(mat), "prix_eur_t"]
    if ser.empty:
        return 0.0
    try:
        val = ser.iloc[0]
        return float(val) if pd.notna(val) else 0.0
    except Exception:
        return 0.0

# =========================
# Onglet 2 : REPROFILAGE SIMPLIFIÉ MULTI-MATÉRIAUX
# =========================
with tab_reprof:
    st.subheader("Reprofilage simplifié (multi-matériaux)")

    # 🧭 UX: BIBLIOTHÈQUE DE MATÉRIAUX éditable (densités t/m³ + épaisseur par défaut cm).
    # Modifiable ici, réutilisée par tous les calculs (Reprofilage et Surfaces).
    with st.container(border=True):
        st.markdown("**📚 Bibliothèque de matériaux** — densités (t/m³), épaisseurs (cm) et **prix unitaire (€/t)**")
        _mat_src = st.session_state.get("materials_df", pd.DataFrame(DEFAULT_MATERIALS))
        # 🧭 UX: garantir la colonne prix (projets/anciennes biblio sans cette colonne)
        if "prix_eur_t" not in _mat_src.columns:
            _mat_src = _mat_src.copy()
            _mat_src["prix_eur_t"] = 0.0
        _mat_edited = st.data_editor(
            _mat_src,
            num_rows="dynamic",
            use_container_width=True,
            key="materials_library_editor",
            column_config={
                "matériau": st.column_config.TextColumn("Matériau", required=True),
                "densité_t_m3": st.column_config.NumberColumn("Densité (t/m³)", min_value=0.0, step=0.05, format="%.2f"),
                "épaisseur_cm": st.column_config.NumberColumn("Épaisseur défaut (cm)", min_value=0.0, step=0.5, format="%.2f"),
                "prix_eur_t": st.column_config.NumberColumn("Prix unitaire (€/t)", min_value=0.0, step=1.0, format="%.2f"),
            },
        )
        # Persister les modifications (ajout/suppression de lignes, densités, épaisseurs)
        try:
            if not _mat_edited.equals(_mat_src):
                st.session_state["materials_df"] = _mat_edited.reset_index(drop=True)
        except Exception:
            st.session_state["materials_df"] = pd.DataFrame(_mat_edited).reset_index(drop=True)
        if st.button("↺ Réinitialiser la bibliothèque par défaut", key="btn_reset_materials"):
            st.session_state["materials_df"] = pd.DataFrame(DEFAULT_MATERIALS)
            st.rerun()
        st.caption("Ajoute/supprime des lignes et ajuste les densités : tout le module Reprofilage (et les Surfaces) utilise cette bibliothèque.")

    # 1) Base de calcul
    base_reprof = st.radio(
        "Base de calcul",
        ["Toute la voirie", "Par élément"],
        horizontal=True,
        key=f"base_reprof_simple__{seg_key}"
    )

    # 2) Source des surfaces (élément + surface_m2)
    reprof_src = _ensure_surfaces_source(base_reprof, recap_elements, surface_totale_voirie).copy()

    # 3) Matériaux disponibles (densités & épaisseurs par défaut)
    materials_df = st.session_state.get("materials_df", pd.DataFrame(DEFAULT_MATERIALS))
    if materials_df.empty:
        st.warning("⚠️ Aucun matériau défini. Ajoute des matériaux avec une densité (t/m³) et une épaisseur par défaut (cm).")
        st.stop()

    mat_opts = materials_df["matériau"].astype(str).tolist()
    rows = []

    # >>> NOUVEAU : mémo local pour reprise "depuis l’élément au‑dessus"
    prev_el_name = None
    prev_el_mats = None
    # Liste d’ordre d’affichage des éléments (utile pour copie depuis un autre élément)
    all_elements_order = reprof_src["élément"].astype(str).tolist()

    # 4) Pour chaque ligne (élément ou globale), N matériaux
    for _, row in reprof_src.iterrows():
        el = str(row["élément"])
        surf = float(row["surface_m2"]) if pd.notna(row["surface_m2"]) else 0.0
        st.markdown(f"### {el} — {surf:,.0f} m²".replace(",", " "))

        # Clé de session dépendant du segment et de l’élément
        key_prefix = f"{seg_key}__{el}"
        list_key = f"mats_{key_prefix}"

        # Initialisation de la liste dynamique des matériaux de cet élément
        if list_key not in st.session_state or not isinstance(st.session_state[list_key], list) or len(st.session_state[list_key]) == 0:
            default_mat = mat_opts[0]
            st.session_state[list_key] = [{
                "mat": default_mat,
                "ep": float(_safe_default_thickness(default_mat, materials_df))
            }]

        mats_list = st.session_state[list_key]

        # --- NOUVEAU : Reprendre depuis l’élément au-dessus (même logique que Rabotage)
        if base_reprof == "Par élément" and prev_el_name is not None:
            with st.expander("Reprendre les matériaux depuis l’élément au‑dessus", expanded=False):
                if st.button(
                    f"⬇️ Copier depuis {prev_el_name}",
                    key=f"reprof_copy_prev_{key_prefix}",
                    use_container_width=True,
                ):
                    # Remplace par une copie profonde de la liste précédente
                    st.session_state[list_key] = [
                        {"mat": str(p.get("mat")), "ep": float(p.get("ep", 0.0))}
                        for p in (prev_el_mats or [])
                    ] or [{
                        "mat": mat_opts[0],
                        "ep": float(_safe_default_thickness(mat_opts[0], materials_df))
                    }]
                    st.rerun()

        # --- NOUVEAU : Copier depuis un autre élément (ex. VL -> VR)
        if base_reprof == "Par élément" and len(all_elements_order) > 1:
            with st.expander("Copier depuis un autre élément", expanded=False):
                others = [e for e in all_elements_order if e != el]
                src_choice = st.selectbox(
                    "Élément source",
                    options=others,
                    key=f"reprof_src_{key_prefix}",
                )
                if st.button(
                    "Copier ici",
                    key=f"reprof_copy_from_{key_prefix}",
                    use_container_width=True,
                ):
                    src_key = f"mats_{seg_key}__{src_choice}"
                    src_list = st.session_state.get(src_key, [])
                    if src_list:
                        st.session_state[list_key] = [
                            {"mat": str(p.get("mat")), "ep": float(p.get("ep", 0.0))}
                            for p in src_list
                        ]
                        st.rerun()
                    else:
                        st.warning(f"Aucun matériau défini pour {src_choice}.")

        # Lignes Matériau × Épaisseur
        for idx, item in enumerate(list(mats_list)):
            col1, col2, col3, col4 = st.columns([2, 2, 2, 1], vertical_alignment="center")

            with col1:
                # Sélection du matériau (densité liée automatiquement)
                try:
                    sel_idx = mat_opts.index(str(item.get("mat", mat_opts[0])))
                except ValueError:
                    sel_idx = 0
                mat = st.selectbox(
                    f"Matériau {idx+1}",
                    mat_opts,
                    index=sel_idx,
                    key=f"mat_{key_prefix}_{idx}"
                )

            with col2:
                dens = float(_safe_default_density(mat, materials_df))
                st.write(f"Densité : **{dens:.2f} t/m³**")

            with col3:
                default_ep = float(_safe_default_thickness(mat, materials_df))
                ep = st.number_input(
                    f"Épaisseur {idx+1} (cm)",
                    min_value=0.0,
                    step=0.5,
                    value=float(item.get("ep", default_ep)),
                    key=f"ep_{key_prefix}_{idx}"
                )

            with col4:
                # Suppression de la ligne
                if st.button("❌", key=f"del_{key_prefix}_{idx}"):
                    mats_list.pop(idx)
                    st.rerun()

            # Mise à jour de l'état
            mats_list[idx]["mat"] = mat
            mats_list[idx]["ep"] = ep

            # Calculs (volume & tonnage pour ce matériau)
            vol = surf * (ep / 100.0)   # m³ = m² × (cm/100)
            ton = vol * dens            # t = m³ × densité (t/m³)

            rows.append({
                "élément": el,
                "matériau": mat,
                "surface_m2": surf,
                "épaisseur_cm": ep,
                "densité_t_m3": dens,
                "volume_m3": vol,
                "tonnage_t": ton
            })

        # Boutons d’action par élément
        c_add, c_reset = st.columns([1, 1])
        with c_add:
            if st.button(f"+ Ajouter un matériau pour {el}", key=f"add_{key_prefix}"):
                default_mat = mat_opts[0]
                mats_list.append({
                    "mat": default_mat,
                    "ep": float(_safe_default_thickness(default_mat, materials_df))
                })
                st.rerun()

        with c_reset:
            if st.button(f"⟲ Réinitialiser {el}", key=f"reset_{key_prefix}"):
                default_mat = mat_opts[0]
                st.session_state[list_key] = [{
                    "mat": default_mat,
                    "ep": float(_safe_default_thickness(default_mat, materials_df))
                }]
                st.rerun()

        # >>> NOUVEAU : mémoriser l’élément courant comme "précédent" pour le suivant
        prev_el_name = el
        prev_el_mats = [
            {"mat": str(p.get("mat")), "ep": float(p.get("ep", 0.0))}
            for p in (st.session_state[list_key] or [])
        ]

    # 5) Résultats & Exports
    df_calc = pd.DataFrame(rows)
    if not df_calc.empty:
        # Colonnes affichées dans l’ordre
        view_cols = ["élément", "matériau", "surface_m2", "densité_t_m3", "épaisseur_cm", "volume_m3", "tonnage_t"]

        st.markdown("### ✅ Détail des calculs")
        st.dataframe(_round_df0(df_calc[view_cols], _DEC_EXCLUDE), use_container_width=True)
        # 🧭 UX: ajoute le détail reprofilage au classeur Excel global
        st.session_state.setdefault("export_bundle", {})["Reprofilage"] = df_calc[view_cols]

        # Totaux globaux
        vol_total = float(df_calc["volume_m3"].sum())
        ton_total = float(df_calc["tonnage_t"].sum())
        k1, k2 = st.columns(2)
        with k1:
            st.metric("Volume total reprofilage (m³)", f"{vol_total:,.0f}".replace(",", " "))
        with k2:
            st.metric("Tonnage total (t)", f"{ton_total:,.0f}".replace(",", " "))

        # Totaux par élément
        st.markdown("#### Totaux par élément")
        tot_el = (
            df_calc.groupby("élément", as_index=False)[["volume_m3", "tonnage_t"]]
            .sum()
            .sort_values("tonnage_t", ascending=False)
        )
        st.dataframe(_round_df0(tot_el, _DEC_EXCLUDE), use_container_width=True)

        # Totaux par matériau
        st.markdown("#### Totaux par matériau")
        tot_mat = (
            df_calc.groupby("matériau", as_index=False)[["volume_m3", "tonnage_t"]]
            .sum()
            .sort_values("tonnage_t", ascending=False)
        )
        st.dataframe(_round_df0(tot_mat, _DEC_EXCLUDE), use_container_width=True)

        # Exports CSV
        _suf = _export_suffix(route, cote, pr_start, pr_end)
        st.download_button(
            "Télécharger le détail (CSV)",
            data=df_calc[view_cols].to_csv(index=False).encode("utf-8"),
            file_name=f"reprofilage_detail_{_suf}.csv",
            mime="text/csv"
        )
        st.download_button(
            "Télécharger totaux par élément (CSV)",
            data=tot_el.to_csv(index=False).encode("utf-8"),
            file_name=f"reprofilage_totaux_par_element_{_suf}.csv",
            mime="text/csv"
        )
        st.download_button(
            "Télécharger totaux par matériau (CSV)",
            data=tot_mat.to_csv(index=False).encode("utf-8"),
            file_name=f"reprofilage_totaux_par_materiau_{_suf}.csv",
            mime="text/csv"
        )
    else:
        st.info("Aucune saisie de matériaux/épaisseurs n’a encore été effectuée pour cette sélection.")



# =========================
# Onglet 3 : SURFACE — créer/nommer des surfaces et calculs simples
# =========================
with tab_surface:
    st.subheader("Surfaces dessinées (polygones)")
    surfs = st.session_state.get("surfaces", {}).get(seg_key, [])
    if not surfs:
        st.info("Aucune surface n'a encore été ajoutée. Dessine un polygone sur la carte puis ajoute-le via le bouton dans la colonne de droite.")
    else:
        rows = []
        _all_surf_mat = []  # 🧭 UX: tonnage par matériau des surfaces (pour le module Coût)
        for i, s in enumerate(surfs):
            st.markdown(f"### {s.get('name','Surface')} - {s.get('area_m2',0):.0f} m²")
            # Renommer
            s['name'] = st.text_input("Nom", value=s.get('name','Surface'), key=f"surf_name_edit_{seg_key}_{i}")
            # Rabotage simple (hauteur unique)
            h = st.number_input("Hauteur de rabotage (cm)", min_value=0.0, step=0.5,
                                value=float(s.get('rabot_h_cm', 0.0)),
                                key=f"surf_rabot_h_{seg_key}_{i}")
            s['rabot_h_cm'] = float(h)
            vol_rabot = float(s.get('area_m2',0.0)) * (h / 100.0)

            # Reprofilage (liste matériaux avec densité par défaut)
            materials_df = st.session_state.get("materials_df", pd.DataFrame(DEFAULT_MATERIALS))
            mat_opts = materials_df["matériau"].astype(str).tolist() if not materials_df.empty else []
            list_key = f"surf_mats_{seg_key}_{i}"
            if list_key not in st.session_state:
                st.session_state[list_key] = s.get('mats', []) or (
                    [{"mat": mat_opts[0], "ep": float(_safe_default_thickness(mat_opts[0], materials_df))}] if mat_opts else []
                )
            mats_list = st.session_state[list_key]

            local_rows = []
            to_del = None
            for j, item in enumerate(list(mats_list)):
                col1, col2, col3, col4 = st.columns([2, 2, 2, 1], vertical_alignment='center')
                with col1:
                    try:
                        idx = mat_opts.index(str(item.get('mat', mat_opts[0] if mat_opts else '')))
                    except ValueError:
                        idx = 0
                    mat = st.selectbox(f"Matériau {j+1}", mat_opts, index=idx if mat_opts else 0,
                                       key=f"surf_mat_{seg_key}_{i}_{j}")
                with col2:
                    dens = float(_safe_default_density(mat, materials_df)) if mat_opts else 0.0
                    st.write(f"Densité : **{dens:.2f} t/m³**")
                with col3:
                    ep = st.number_input(f"Épaisseur {j+1} (cm)", min_value=0.0, step=0.5,
                                         value=float(item.get('ep', _safe_default_thickness(mat, materials_df))),
                                         key=f"surf_ep_{seg_key}_{i}_{j}")
                with col4:
                    if st.button("❌", key=f"surf_del_{seg_key}_{i}_{j}"):
                        to_del = j
                # maj
                if j < len(mats_list):
                    mats_list[j] = {"mat": mat, "ep": ep}
                vol = float(s.get('area_m2',0.0)) * (float(ep) / 100.0)
                ton = vol * dens
                local_rows.append({"surface": s.get('name','Surface'), "matériau": mat, "épaisseur_cm": ep,
                                   "volume_m3": vol, "tonnage_t": ton})

            if to_del is not None:
                mats_list.pop(to_del); st.rerun()

            c1, c2 = st.columns(2)
            with c1:
                if st.button("+ Ajouter un matériau", key=f"surf_addmat_{seg_key}_{i}"):
                    if mat_opts:
                        mats_list.append({"mat": mat_opts[0], "ep": float(_safe_default_thickness(mat_opts[0], materials_df))})
                        st.rerun()
            with c2:
                if st.button("⟲ Réinitialiser matériaux", key=f"surf_resetmat_{seg_key}_{i}"):
                    st.session_state[list_key] = [{"mat": mat_opts[0], "ep": float(_safe_default_thickness(mat_opts[0], materials_df))}] if mat_opts else []
                    st.rerun()

            # Persister dans l'objet surface
            s['mats'] = mats_list
            _all_surf_mat.extend(local_rows)  # 🧭 UX: cumul tonnages matériaux des surfaces
            # Récap surface
            if local_rows:
                df_loc = pd.DataFrame(local_rows)
                st.dataframe(_round_df0(df_loc, _DEC_EXCLUDE), use_container_width=True)
                vol_reprof = float(df_loc['volume_m3'].sum())
                ton_reprof = float(df_loc['tonnage_t'].sum())
            else:
                vol_reprof, ton_reprof = 0.0, 0.0
            rows.append({
                "surface": s.get('name','Surface'),
                "aire_m2": float(s.get('area_m2',0.0)),
                "rabot_h_cm": float(h),
                "vol_rabot_m3": float(vol_rabot),
                "vol_reprof_m3": float(vol_reprof),
                "tonnage_t": float(ton_reprof),
            })

            if st.button(f"🗑️ Supprimer {s.get('name','Surface')}", key=f"surf_delete_{seg_key}_{i}"):
                surfs.pop(i)
                st.session_state.get("surfaces", {}).setdefault(seg_key, surfs)
                st.rerun()

        # Summary global + export
        st.session_state.get("surfaces", {}).setdefault(seg_key, surfs)
        df_summary = pd.DataFrame(rows)
        st.markdown("#### ✅ Récapitulatif Surfaces")
        st.dataframe(_round_df0(df_summary, _DEC_EXCLUDE), use_container_width=True)
        # 🧭 UX: ajoute le récap surfaces au classeur Excel global
        st.session_state.setdefault("export_bundle", {})["Surfaces"] = df_summary
        # 🧭 UX: tonnage par matériau des surfaces, pour le module Coût
        st.session_state.setdefault("export_bundle", {})["Surfaces_matériaux"] = pd.DataFrame(_all_surf_mat)
        k1, k2, k3 = st.columns(3)
        with k1: st.metric("Aire totale (m²)", f"{df_summary['aire_m2'].sum():.0f}")
        with k2: st.metric("Volume rabotage total (m³)", f"{df_summary['vol_rabot_m3'].sum():,.0f}".replace(',', ' '))
        with k3: st.metric("Tonnage reprofilage total (t)", f"{df_summary['tonnage_t'].sum():,.0f}".replace(',', ' '))
        st.download_button(
            "Télécharger le récap surfaces (CSV)",
            data=df_summary.to_csv(index=False).encode('utf-8'),
            file_name=f"surfaces_recap_{_export_suffix(route, cote, pr_start, pr_end)}.csv",
            mime="text/csv",
        )

# =========================
# 🧭 UX: Onglet 4 — COÛT (€) : tonnage par matériau × prix unitaire (depuis la bibliothèque)
# =========================
with tab_cout:
    st.subheader("Estimation du coût (€)")
    _materials_df = st.session_state.get("materials_df", pd.DataFrame(DEFAULT_MATERIALS))
    _eb_cost = st.session_state.get("export_bundle", {})

    # Sources de tonnage par matériau : Reprofilage (+ Surfaces dessinées en option)
    _inc_surf = st.checkbox("Inclure le tonnage des surfaces dessinées", value=True, key="cost_include_surfaces")
    _src_frames = []
    _rep = _eb_cost.get("Reprofilage")
    if _rep is not None and not _rep.empty and {"matériau", "tonnage_t"}.issubset(_rep.columns):
        _src_frames.append(_rep[["matériau", "tonnage_t"]])
    if _inc_surf:
        _sm = _eb_cost.get("Surfaces_matériaux")
        if _sm is not None and not _sm.empty and {"matériau", "tonnage_t"}.issubset(_sm.columns):
            _src_frames.append(_sm[["matériau", "tonnage_t"]])

    if not _src_frames:
        st.info("Renseigne d'abord le **Reprofilage** (et/ou les **Surfaces**) pour estimer un coût.")
    else:
        _ton = pd.concat(_src_frames, ignore_index=True)
        _ton["tonnage_t"] = pd.to_numeric(_ton["tonnage_t"], errors="coerce").fillna(0.0)
        cout_df = _ton.groupby("matériau", as_index=False)["tonnage_t"].sum()
        cout_df["prix_eur_t"] = cout_df["matériau"].map(lambda m: _safe_default_price(m, _materials_df))
        cout_df["coût_eur"] = cout_df["tonnage_t"] * cout_df["prix_eur_t"]
        cout_df = cout_df.sort_values("coût_eur", ascending=False).reset_index(drop=True)

        st.markdown("#### Coût par matériau")
        st.dataframe(
            _round_df0(cout_df, exclude=["prix_eur_t"]),
            use_container_width=True,
            column_config={
                "tonnage_t": st.column_config.NumberColumn("Tonnage (t)", format="%d"),
                "prix_eur_t": st.column_config.NumberColumn("Prix unitaire (€/t)", format="%.2f"),
                "coût_eur": st.column_config.NumberColumn("Coût (€)", format="%d"),
            },
        )

        _tot_ton = float(cout_df["tonnage_t"].sum())
        _tot_cout = float(cout_df["coût_eur"].sum())
        kc1, kc2 = st.columns(2)
        with kc1:
            st.metric("Tonnage total (t)", f"{_tot_ton:,.0f}".replace(",", " "))
        with kc2:
            st.metric("💶 Coût total estimé (€)", f"{_tot_cout:,.0f}".replace(",", " "))

        if (cout_df["prix_eur_t"] == 0).any():
            st.warning("Certains matériaux ont un prix unitaire à 0 €/t. Renseigne-les dans la "
                       "**📚 Bibliothèque de matériaux** (onglet Reprofilage) pour un coût complet.")
        st.caption("Les prix unitaires (€/t) se paramètrent dans la bibliothèque de matériaux. "
                   "Le coût combine le reprofilage" + (" et les surfaces dessinées." if _inc_surf else "."))

        # Alimente le rapport Excel + export CSV dédié
        st.session_state.setdefault("export_bundle", {})["Coût"] = cout_df
        st.download_button(
            "Télécharger l'estimation de coût (CSV)",
            data=cout_df.to_csv(index=False).encode("utf-8"),
            file_name=f"cout_estimation_{_export_suffix(route, cote, pr_start, pr_end)}.csv",
            mime="text/csv",
        )


# =========================
# Calcul batch (optionnel) — inchangé (avec facteur de courbure)
# =========================
with st.expander("Calculer tous les intervalles consécutifs de ce côté (batch)", expanded=False):
    run_batch = st.checkbox("Activer le calcul batch")
    if run_batch:
        results = []
        pr_vals = subset["pr"].dropna().unique().tolist()
        pr_vals = sorted(pr_vals)
        for a, b in zip(pr_vals[:-1], pr_vals[1:]):
            row_a = subset[subset["pr"] == a].iloc[0]
            row_b = subset[subset["pr"] == b].iloc[0]
            # distance pour le batch: chainage_m si dispo, sinon droite PR→PR
            if pd.notna(row_a.get("chainage_m", np.nan)) and pd.notna(row_b.get("chainage_m", np.nan)):
                d = float(row_b["chainage_m"] - row_a["chainage_m"])
            else:
                d = planimetric_distance_l93(
                    [(float(row_a["x"]), float(row_a["y"])),
                     (float(row_b["x"]), float(row_b["y"]))]
                )
            d = max(d * float(curvature_factor), 0.0)
            widths_pair = apply_overrides(widths_applied, overrides, route, cote, float(a), float(b))
            _, total_pair = compute_areas(d, widths_pair, profile_mix, included_elements)
            results.append(
                {
                    "route": route, "côté": cote, "PR_début": a, "PR_fin": b,
                    "distance_m": d, "surface_totale_m²": total_pair
                }
            )
        res_df_fr = pd.DataFrame(results)
        st.dataframe(_round_df0(res_df_fr, _DEC_EXCLUDE), use_container_width=True)
        # Export CSV
        csv_bytes = res_df_fr.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Télécharger le récap batch (CSV)",
            data=csv_bytes, file_name="batch_surfaces.csv",
            mime="text/csv",
        )

# =========================
# Aide rapide
# =========================
with st.expander("Aide rapide", expanded=False):
    st.markdown(
        """
- **Import** : CSV (`;` et `,` décimale) ou Excel. Colonnes requises : `route`, `cote`, `pr`, `x`, `y`.
  Colonne optionnelle : `chainage_m` (dans ton CSV, `cumul` → `chainage_m` automatiquement).
- **Distances** :
  - **Segment édité** : distance planimétrique de la/les polyligne(s) éditée(s).
  - **Chainage** : utilise `chainage_m` si présent, sinon `1000 m` par PR.
  - **Droite PR→PR** : distance droite entre les 2 PR.
  - **PR × 1000 (fixe)** : impose `1000 m` par PR, même si `chainage_m` existe.
  - **Fixe** : valeur imposée (par défaut `1000 × ΔPR`, modifiable).
- **Sous-segments (version dessin)** : dessine une ligne, choisis un profil (unique) et clique **“➕ Ajouter comme sous‑segment”**.
  Chaque sous‑segment garde sa couleur (profil dominant) **et affiche sa distance** dans le tableau et sur la carte.
  Les noms sont incrémentés par **profil dominant** (ex. `3_voies_1`, `3_voies_2`, ...).
- **Éléments** : préréglages “Voies (enrobé)”, “Tout”, ou sélection personnalisée.
- **Largeurs** : modifie les largeurs par élément ; **overrides** (route, cote, pr_start, pr_end, element, largeur_m) s’appliquent au **tronçon sélectionné**.
- **Récapitulatif global** : affiche la **surface totale de la voirie** et la **surface cumulée par élément** (tous sous‑segments confondus).
"""
    )

# =========================
# 🧭 UX: RAPPORT EXCEL propre à 4 onglets (Surface général · Rabotage · Reprofilage · Surface)
# =========================
st.markdown("---")
st.markdown("#### 📊 Rapport Excel (5 onglets — charte VINCI)")
_eb = st.session_state.get("export_bundle", {})
try:
    _report_ctx = {
        "route": route, "cote": cote, "pr_start": pr_start, "pr_end": pr_end,
        "distance_m": float(distance_display_m),
        "methode": dist_method,
        "profil_dominant": (dominant_profile_name(profiles_selected, percents) or "mix").replace("_", " "),
        "surface_totale_voirie": float(surface_totale_voirie),
    }
    _report_bytes = _build_report_xlsx(
        _report_ctx,
        areas_df_fr,
        recap_elements,
        _eb.get("Rabotage"),
        _eb.get("Reprofilage"),
        _eb.get("Surfaces"),
        _eb.get("Coût"),
    )
    st.download_button(
        "📊 Télécharger le rapport Excel (5 onglets)",
        data=_report_bytes,
        file_name=f"rapport_estimation_{_export_suffix(route, cote, pr_start, pr_end)}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(
        "Onglets : **Surface général** · **Rabotage** · **Reprofilage** · **Surface** · **Coût (€)** — "
        "charte VINCI, totaux en évidence et détails par voie / matériau / profil. Renseigne les onglets "
        "Rabotage/Reprofilage/Surface/Coût pour remplir les feuilles. Les exports CSV par section restent dispo."
    )
except Exception as _re:
    st.warning(f"Rapport Excel indisponible pour le moment : {_re}")
